#!/usr/bin/env python3
"""Read the team's daily Excel report and reconcile it against the machine trails.

    python3 importers/daily_report.py --file inbox/daily.xlsx \\
        --events exports/jira-PRJ.ndjson --events exports/aio-PRJ.ndjson \\
        --out mapping.json

Column spec: ``docs/IMPORT-SPEC.md``.

**This importer emits no events, deliberately.** The daily report is
self-reported, written after the fact, by the people being measured. Counting
from it would rank people by how carefully they write reports -- someone who
works hard and logs little looks worse than someone who works less and logs
everything, and nothing in the output would reveal that it happened. So it
produces two things instead:

1. **A mapping table.** AIO test cases carry no Jira key (0 of 4,248 measured),
   which blocks Automation Lead Time and any split of coverage by priority.
   Where the report says a person touched a Jira item and a test case on the
   same day, *and both machine trails agree that something happened*, that is
   evidence for a link -- built without hand-editing 4,248 cases.

2. **Trail completeness.** What share of self-reported work left a
   machine-readable trace. Disagreement is a finding either way: work that is
   not reaching Jira, or a report that overstates. Both are worth surfacing and
   neither should be smoothed over.

The ``note`` column is read and **discarded**. It is free text written by a
person, which is content, and CONTRACT.md §1.1 does not admit exceptions for
content that happens to be convenient.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date as date_type, timedelta
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if os.path.join(_ROOT, "pollers") not in sys.path:
    sys.path.insert(0, os.path.join(_ROOT, "pollers"))

import common  # noqa: E402

REQUIRED = ("date", "person", "jira_key", "activity", "ai_used")
OPTIONAL = ("test_case_key", "ai_agent", "hours", "note")

ACTIVITIES = frozenset(
    {"design_case", "automate", "execute", "fix_defect", "review", "other"})

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
JIRA_KEY_RE = re.compile(r"^[A-Z][A-Z0-9]+-\d+$")
CASE_KEY_RE = re.compile(r"^[A-Z][A-Z0-9]+-TC-\d+$")
TRUE = frozenset({"yes", "y", "true", "1"})
FALSE = frozenset({"no", "n", "false", "0"})


class Row:
    """One reported piece of work. ``note`` is never kept."""

    __slots__ = ("number", "date", "person_hash", "jira_key", "test_case_key",
                 "activity", "ai_used", "ai_agent", "hours")

    def __init__(self, number, date, person_hash, jira_key, test_case_key,
                 activity, ai_used, ai_agent, hours):
        self.number = number
        self.date = date
        self.person_hash = person_hash
        self.jira_key = jira_key
        self.test_case_key = test_case_key
        self.activity = activity
        self.ai_used = ai_used
        self.ai_agent = ai_agent
        self.hours = hours

    def as_dict(self) -> Dict[str, Any]:
        return {slot: getattr(self, slot) for slot in self.__slots__}


def read_workbook(path: str) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency guard
        sys.exit("daily_report.py needs openpyxl:  python3 -m pip install openpyxl")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def cell(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        # Excel hands back a datetime when the cell was formatted as a date.
        # The spec asks for text precisely because this conversion is
        # locale-dependent, but accept it rather than reject a whole file.
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def parse_rows(table: Sequence[Sequence[Any]], salt: Optional[str]
               ) -> Tuple[List[Row], List[Dict[str, Any]]]:
    """Parse, or report. A row is never silently dropped or guessed at.

    An unparseable row is returned as a problem carrying its row number, so it
    can be fixed at the source. Coercing it into a guess would put invented data
    in the mapping table with the same confidence as the real thing.
    """
    problems: List[Dict[str, Any]] = []
    if not table:
        return [], [{"row": 0, "problem": "the sheet is empty"}]

    header = [cell(c).lower() for c in table[0]]
    missing = [name for name in REQUIRED if name not in header]
    if missing:
        return [], [{"row": 1, "problem": "missing required column(s): {}".format(
            ", ".join(missing))}]
    index = {name: header.index(name) for name in header if name}

    def get(row, name):
        position = index.get(name)
        if position is None or position >= len(row):
            return ""
        return cell(row[position])

    rows: List[Row] = []
    for number, raw in enumerate(table[1:], start=2):
        if not any(cell(c) for c in raw):
            continue  # a blank spacer row is not an error

        faults = []
        date = get(raw, "date")
        if not DATE_RE.match(date):
            faults.append("date {!r} is not YYYY-MM-DD".format(date))

        person = get(raw, "person")
        if "@" not in person:
            faults.append("person {!r} is not an email".format(person))

        jira_key = get(raw, "jira_key").upper()
        if not JIRA_KEY_RE.match(jira_key):
            faults.append("jira_key {!r} is not a key".format(jira_key))

        case_key = get(raw, "test_case_key").upper()
        if case_key and not CASE_KEY_RE.match(case_key):
            faults.append("test_case_key {!r} is not a key".format(case_key))

        activity = get(raw, "activity").lower()
        if activity not in ACTIVITIES:
            faults.append("activity {!r} is not one of: {}".format(
                activity, ", ".join(sorted(ACTIVITIES))))

        ai_raw = get(raw, "ai_used").lower()
        if ai_raw in TRUE:
            ai_used = True
        elif ai_raw in FALSE:
            ai_used = False
        else:
            ai_used = None
            faults.append("ai_used {!r} is not yes or no".format(ai_raw))

        hours_raw = get(raw, "hours")
        hours = None
        if hours_raw:
            try:
                hours = float(hours_raw)
            except ValueError:
                faults.append("hours {!r} is not a number".format(hours_raw))

        if faults:
            # The row number is the point of this: it is what makes the problem
            # fixable at the source rather than an anonymous count.
            problems.append({"row": number, "problem": "; ".join(faults)})
            continue

        rows.append(Row(
            number=number, date=date,
            person_hash=common.hash_email(person, salt),
            jira_key=jira_key, test_case_key=case_key or None,
            activity=activity, ai_used=ai_used,
            ai_agent=get(raw, "ai_agent") or None, hours=hours,
        ))
    return rows, problems


def index_events(paths: Iterable[str]) -> Dict[str, Dict[str, set]]:
    """Which Jira keys and test case keys show activity on which day."""
    jira: Dict[str, set] = defaultdict(set)
    cases: Dict[str, set] = defaultdict(set)
    for path in paths:
        with open(path, "r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                event = json.loads(line)
                day = (event.get("event_time") or "")[:10]
                if not day:
                    continue
                attributes = event.get("attributes") or {}
                key = (event.get("context") or {}).get("jira_issue_key") \
                    or attributes.get("jira_issue_key")
                if key:
                    jira[key].add(day)
                case = attributes.get("test_case_key")
                if case:
                    cases[case].add(day)
    return {"jira": jira, "cases": cases}


#: A report is written at the end of a day or the morning after, and a Jira
#: transition can land either side of the work it records. Matching the exact
#: calendar day makes trail completeness read near zero for reasons that have
#: nothing to do with whether the work left a trace -- and a metric that always
#: reads zero gets dismissed rather than investigated.
DEFAULT_DAY_TOLERANCE = 1


def within(day: str, days: Iterable[str], tolerance: int) -> bool:
    if day in days:
        return True
    if tolerance <= 0:
        return False
    try:
        target = date_type.fromisoformat(day)
    except ValueError:
        return False
    for offset in range(1, tolerance + 1):
        for shifted in (target - timedelta(days=offset),
                        target + timedelta(days=offset)):
            if shifted.isoformat() in days:
                return True
    return False


def reconcile(rows: Sequence[Row], trails: Dict[str, Dict[str, set]],
              tolerance: int = DEFAULT_DAY_TOLERANCE) -> Dict[str, Any]:
    """Cross the report against the machine trails.

    A pair is only proposed when both sides show activity within ``tolerance``
    days of the reported one. One-sided agreement is recorded at lower
    confidence rather than promoted: the whole point of a mapping table is that
    a later metric can trust it, and a table that mixes corroborated and
    uncorroborated rows at one confidence is not trustworthy at any.
    """
    jira_trail = trails.get("jira", {})
    case_trail = trails.get("cases", {})

    pairs: Dict[Tuple[str, str], Dict[str, Any]] = {}
    corroborated = 0
    for row in rows:
        in_jira = within(row.date, jira_trail.get(row.jira_key, set()), tolerance)
        if in_jira:
            corroborated += 1
        if not row.test_case_key:
            continue
        in_aio = within(row.date, case_trail.get(row.test_case_key, set()), tolerance)
        key = (row.jira_key, row.test_case_key)
        entry = pairs.setdefault(key, {
            "jira_key": row.jira_key, "test_case_key": row.test_case_key,
            "observations": 0, "both_trails_agree": 0, "days": [],
        })
        entry["observations"] += 1
        if in_jira and in_aio:
            entry["both_trails_agree"] += 1
        if row.date not in entry["days"]:
            entry["days"].append(row.date)

    mappings = []
    for entry in pairs.values():
        agreed = entry["both_trails_agree"]
        mappings.append({
            **entry,
            "days": sorted(entry["days"]),
            "confidence": "corroborated" if agreed else "reported_only",
        })
    mappings.sort(key=lambda m: (m["confidence"] != "corroborated",
                                 m["jira_key"], m["test_case_key"]))

    reported_keys = {row.jira_key for row in rows}
    return {
        "day_tolerance": tolerance,
        "mappings": mappings,
        "corroborated_mappings": sum(
            1 for m in mappings if m["confidence"] == "corroborated"),
        "trail_completeness": {
            "rows": len(rows),
            "rows_with_a_jira_trail": corroborated,
            # Reported without a trace is not automatically a bad report -- it
            # is equally often work that never reached Jira. It is a question,
            # and the number exists so somebody asks it.
            "rows_without_a_jira_trail": len(rows) - corroborated,
            "pct": round(100.0 * corroborated / len(rows), 1) if rows else None,
            "distinct_jira_keys_reported": len(reported_keys),
            # Separate from the row count on purpose: a key that exists in Jira
            # but shows nothing near the reported day means the report and the
            # trail disagree about *when*, which is a different problem from
            # work that never reached Jira at all.
            "distinct_jira_keys_with_a_trail": len(
                [k for k in reported_keys if jira_trail.get(k)]),
        },
        "ai_agents_reported": sorted(
            {row.ai_agent for row in rows if row.ai_agent}),
    }


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Reconcile the daily Excel report against the machine trails.")
    parser.add_argument("--file", required=True, help="the daily report .xlsx")
    parser.add_argument("--events", action="append", default=[],
                        help="NDJSON event file to reconcile against (repeatable)")
    parser.add_argument("--salt", default=os.environ.get("TELEMETRY_HASH_SALT"),
                        help="hashing salt; raw emails never reach storage")
    parser.add_argument("--out", help="write the mapping report here")
    parser.add_argument("--day-tolerance", type=int, default=DEFAULT_DAY_TOLERANCE,
                        help="days either side of a reported date that still "
                             "count as corroboration (default: %(default)s)")
    parser.add_argument("--strict", action="store_true",
                        help="exit non-zero if any row failed to parse")
    args = parser.parse_args(argv)

    rows, problems = parse_rows(read_workbook(args.file), args.salt)
    trails = index_events(args.events) if args.events else {"jira": {}, "cases": {}}
    report = reconcile(rows, trails, args.day_tolerance)
    report["source_file"] = os.path.basename(args.file)
    report["rows_parsed"] = len(rows)
    report["rows_rejected"] = len(problems)
    report["problems"] = problems

    if args.out:
        with open(args.out, "w", encoding="utf-8") as handle:
            json.dump(report, handle, indent=2, sort_keys=True)

    print(json.dumps({
        "msg": "daily_report_import_complete",
        "rows_parsed": len(rows),
        "rows_rejected": len(problems),
        "mappings": len(report["mappings"]),
        "corroborated": report["corroborated_mappings"],
        "trail_completeness_pct": report["trail_completeness"]["pct"],
    }, sort_keys=True), file=sys.stderr)
    for problem in problems:
        print("ROW {}: {}".format(problem["row"], problem["problem"]), file=sys.stderr)

    return 1 if (args.strict and problems) else 0


if __name__ == "__main__":
    sys.exit(main())
