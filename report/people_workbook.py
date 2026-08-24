#!/usr/bin/env python3
"""people_workbook.py -- per-person Excel workbook from the telemetry event stream.

Reads the NDJSON emitted by the pollers (CONTRACT.md §3 events) and writes one
.xlsx with a sheet per view, scoped to a named set of people.

    python3 people_workbook.py \
        --person "A Name=<atlassian-account-id>" \
        --person "Another Name=<atlassian-account-id>" \
        --input ../../../reports/exports \
        --since 2026-07-16 \
        --out ../../../reports/ai-work-tracking.xlsx

Account ids are Atlassian ``accountId`` values, the same key Jira, Bitbucket and
AIO all use. Find one with::

    curl -s -u "$JIRA_USERNAME:$JIRA_API_TOKEN" \
        "$JIRA_URL/rest/api/3/user/search?query=<name or email>"

Why a workbook and not the weekly report
----------------------------------------
``weekly_report.py`` is deliberately **team-level and person-blind** -- design
§11.5 keeps a per-person view behind a governance sign-off, and the report
refuses to render one. This tool is the explicit, named exception: a lead
tracking specific engineers they are directly supporting, with each of those
people named on the command line.

It is not a leaderboard generator. It takes an **allow-list** and has no way to
express "top N by output" -- there is no ranking, no sort by volume, and no
query that discovers who to look at. Deciding whom to include is a human
decision made outside this tool, which is the control that keeps it from
becoming a surveillance report.

Any number of people may be passed; two is merely the smallest useful case.

Two rules it keeps from the report, because they are what make the numbers
survive a challenge:

* **Absent is not zero.** A metric with no source events renders ``n/a`` and a
  reason, never ``0``. See the Coverage sheet.
* **Median, not mean**, for every lead time. These distributions are long-tailed
  by construction and a mean hides the tail that actually hurts.

Third parties who appear only as a reviewer or a co-assignee are written as
their opaque account id, never a display name. Only the people named on the
command line get a name in the file.

Requires ``openpyxl``. Everything else is stdlib.
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import statistics
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, Iterable, List, Optional, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.data_source import AxDataSource, StrRef
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guard
    sys.exit("people_workbook.py needs openpyxl:  python3 -m pip install openpyxl")


# --------------------------------------------------------------------------
# loading
# --------------------------------------------------------------------------

def iter_files(inputs: Sequence[str]) -> Iterable[str]:
    for path in inputs:
        if os.path.isdir(path):
            yield from sorted(glob.glob(os.path.join(path, "**", "*.ndjson"),
                                        recursive=True))
        else:
            yield path


def parse_ts(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def load_events(inputs: Sequence[str]) -> tuple[List[dict], Dict[str, int]]:
    events: List[dict] = []
    seen: set[str] = set()
    stats = {"files": 0, "lines": 0, "malformed": 0, "duplicates": 0}
    for path in iter_files(inputs):
        stats["files"] += 1
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                stats["lines"] += 1
                try:
                    event = json.loads(line)
                except json.JSONDecodeError:
                    stats["malformed"] += 1
                    continue
                if not isinstance(event, dict) or "event_type" not in event:
                    stats["malformed"] += 1
                    continue
                eid = event.get("event_id")
                if eid and eid in seen:
                    stats["duplicates"] += 1
                    continue
                if eid:
                    seen.add(eid)
                events.append(event)
    return events, stats


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------

def ms_to_hours(ms: Optional[int]) -> Optional[float]:
    return None if ms is None else round(ms / 3_600_000, 2)


def med(values: Sequence[float]) -> Optional[float]:
    vals = [v for v in values if v is not None]
    return round(statistics.median(vals), 2) if vals else None


def p85(values: Sequence[float]) -> Optional[float]:
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return None
    idx = min(len(vals) - 1, int(round(0.85 * (len(vals) - 1))))
    return round(vals[idx], 2)


def na(value: Any) -> Any:
    """Render an absent measurement as text, never as a zero."""
    return "n/a" if value is None else value


def fmt_list(values: Optional[Iterable[Any]]) -> str:
    return ", ".join(str(v) for v in (values or [])) or ""


# --------------------------------------------------------------------------
# aggregation
# --------------------------------------------------------------------------

class PersonData:
    def __init__(self, name: str, account_id: str):
        self.name = name
        self.account_id = account_id
        self.issues: Dict[str, dict] = {}
        self.transitions: List[dict] = []
        self.prs: Dict[int, dict] = {}
        self.reviews: List[dict] = []
        self.reverts: List[dict] = []
        self.test_runs: List[dict] = []      # executed by this person
        self.test_assigned: List[dict] = []  # assigned to them, not yet run
        self.owned_cases: List[dict] = []    # test cases they own (coverage)


def collect(events: Sequence[dict], people: Dict[str, str],
            since: Optional[datetime]) -> Dict[str, PersonData]:
    data = {name: PersonData(name, aid) for name, aid in people.items()}
    by_id = {aid: name for name, aid in people.items()}

    for event in events:
        when = parse_ts(event.get("event_time"))
        if since and when and when < since:
            continue
        etype = event.get("event_type")
        attrs = event.get("attributes") or {}
        ctx = event.get("context") or {}
        actor = (event.get("actor") or {}).get("person_id")

        if etype == "jira.transition":
            issue = attrs.get("issue") or {}
            # An engineer "owns" an issue if they are assigned it or raised it.
            # The person who merely clicked the transition is recorded too, but
            # separately -- a QA lead moving someone else's ticket to Done is
            # not that lead's delivered work.
            for role_id, role in ((issue.get("assignee_person_id"), "assignee"),
                                  (issue.get("reporter_person_id"), "reporter"),
                                  (actor, "transitioned")):
                name = by_id.get(role_id or "")
                if not name:
                    continue
                person = data[name]
                key = issue.get("issue_key")
                if key:
                    record = person.issues.setdefault(key, {
                        "issue": issue, "roles": set(), "transitions": 0,
                        "attribution": attrs.get("attribution") or {},
                    })
                    record["roles"].add(role)
                    if role == "assignee":
                        record["transitions"] += 1
                if role == "assignee":
                    person.transitions.append({
                        "issue_key": key,
                        "from_status": attrs.get("from_status"),
                        "to_status": attrs.get("to_status"),
                        "status_category": attrs.get("status_category"),
                        "transitioned_at": attrs.get("transitioned_at"),
                        "age_hours": ms_to_hours(attrs.get("age_at_transition_ms")),
                        "synthesised": attrs.get("is_synthesised_creation"),
                    })

        elif etype in ("scm.pr.merged", "scm.pr.declined", "scm.pr.opened"):
            name = by_id.get(actor or "")
            if not name:
                continue
            pr_id = attrs.get("pr_id")
            data[name].prs[pr_id] = {
                "event_type": etype, "attrs": attrs, "ctx": ctx,
                "event_time": event.get("event_time"),
            }

        elif etype == "scm.pr.reviewed":
            reviewer = attrs.get("reviewer_person_id") or actor
            name = by_id.get(reviewer or "")
            if not name:
                continue
            data[name].reviews.append({"attrs": attrs, "ctx": ctx})

        elif etype == "scm.revert":
            name = by_id.get(actor or "")
            if not name:
                continue
            data[name].reverts.append({"attrs": attrs, "ctx": ctx})

        elif etype == "test.case.snapshot":
            # Coverage is a property of the inventory, not of a run: a case
            # nobody has executed emits no run event, and those are exactly the
            # un-automated ones the denominator needs.
            name = by_id.get(actor or "")
            if name:
                data[name].owned_cases.append({"attrs": attrs})

        elif etype == "test.run.completed":
            # Executed-by and assigned-to are kept apart on purpose. Running a
            # test case is work; being the standing assignee on a case someone
            # else executed, or on one nobody has executed, is not.
            executed_by = attrs.get("executed_by_person_id")
            assigned_to = attrs.get("assigned_to_person_id")
            executor = by_id.get(executed_by or "")
            if executor:
                data[executor].test_runs.append({"attrs": attrs})
            owner = by_id.get(assigned_to or "")
            if owner and owner != executor:
                data[owner].test_assigned.append({"attrs": attrs})

    return data


TEST_EXECUTED = frozenset({"passed", "failed", "blocked", "skipped"})


def coverage_totals(person: PersonData) -> Dict[str, int]:
    """Automation coverage over the test cases this person owns.

    The denominator is cases with a **known** automation status. An unset field
    is not "not automated" -- about 42% of this estate has never had it set --
    and folding those in would measure how diligently the field is filled in
    rather than how much is automated.
    """
    counts: Counter = Counter()
    for record in person.owned_cases:
        attrs = record["attrs"]
        if attrs.get("is_archived"):
            counts["archived"] += 1
            continue
        status = (attrs.get("automation_status") or "").strip().lower()
        if not status:
            counts["status_unset"] += 1
        elif status == "automated":
            counts["automated"] += 1
        elif status.startswith("to be"):
            counts["to_be_automated"] += 1
        else:
            counts["status_other"] += 1
        counts["owned"] += 1
    counts["known"] = counts["automated"] + counts["to_be_automated"]
    return counts


def test_totals(person: PersonData) -> Dict[str, int]:
    """Counts by status category, plus the executed denominator.

    ``not_run`` is excluded from ``executed`` deliberately: AIO seeds every new
    cycle with one row per test case at "Not Run", so counting them would turn
    cycle planning into apparent test activity, and would put a large
    denominator under a pass rate that nobody earned.
    """
    counts: Counter = Counter()
    for run in person.test_runs:
        counts[run["attrs"].get("status_category") or "other"] += 1
    counts["executed"] = sum(counts[c] for c in TEST_EXECUTED)
    return counts


# --------------------------------------------------------------------------
# sheet writing
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=12)
NOTE_FONT = Font(italic=True, color="666666")


def write_sheet(wb: Workbook, title: str, headers: Sequence[str],
                rows: Sequence[Sequence[Any]], note: str = "") -> None:
    ws = wb.create_sheet(title)
    ws.append(list(headers))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        )
    for idx, header in enumerate(headers, start=1):
        widest = max([len(str(header))] +
                     [len(str(r[idx - 1])) for r in rows[:400]
                      if idx - 1 < len(r) and r[idx - 1] is not None] or [10])
        ws.column_dimensions[get_column_letter(idx)].width = min(46, max(10, widest + 2))
    if note:
        ws.append([])
        ws.append([note])
        ws.cell(row=ws.max_row, column=1).font = NOTE_FONT


def build_summary(wb: Workbook, data: Dict[str, PersonData],
                  window: str, gaps: Sequence[Sequence[str]]) -> None:
    ws = wb.create_sheet("Summary", 0)
    names = list(data)

    ws.append(["AI work tracking", window])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    def metric(label: str, fn, note: str = "", basis=None) -> None:
        """Write one metric row.

        ``basis`` separates the two things that must never look alike: a
        measured zero and an absent measurement. If ``basis(person)`` is falsy
        there was nothing to count from, so the cell reads ``no data``. If it is
        truthy, a zero is a real zero and is printed as ``0``.

        The first real run of this tool found an engineer with 32 merged pull
        requests and zero reviewers on any of them -- a finding about how that
        repository is used. Rendering it as ``n/a`` would have hidden it behind
        an apparent gap in the data.
        """
        cells = []
        for n in names:
            person = data[n]
            if basis is not None and not basis(person):
                cells.append("no data")
            else:
                value = fn(person)
                cells.append("no data" if value is None else value)
        ws.append([label] + cells + [note])

    has_issues = lambda p: bool(p.issues)          # noqa: E731
    has_prs = lambda p: bool(p.prs)                # noqa: E731
    any_scm = lambda p: bool(p.prs or p.reviews or p.reverts)  # noqa: E731

    ws.append(["Metric"] + names + ["Note"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    header_row = ws.max_row

    ws.append(["JIRA"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    metric("Issues assigned", lambda p: sum(1 for r in p.issues.values()
                                            if "assignee" in r["roles"]))
    metric("Issues raised (reporter)", lambda p: sum(1 for r in p.issues.values()
                                                     if "reporter" in r["roles"]))
    metric("Issues resolved", lambda p: sum(
        1 for r in p.issues.values()
        if "assignee" in r["roles"] and r["issue"].get("resolved_at")))
    metric("Bugs assigned", lambda p: sum(
        1 for r in p.issues.values()
        if "assignee" in r["roles"] and r["issue"].get("issue_type") == "Bug"))
    metric("Status transitions", lambda p: len(p.transitions), basis=has_issues)
    metric("Median issue age at transition (h)",
           lambda p: med([t["age_hours"] for t in p.transitions]),
           "Median, not mean -- the distribution is long-tailed", basis=has_issues)
    metric("Issues carrying an AI label", lambda p: sum(
        1 for r in p.issues.values() if r["attribution"].get("has_ai_labels")),
        "AUTH_/PLANNED_/GEN_/REVIEW_BY_COPILOT, may sit on the delivery ticket (AR-3)",
        basis=has_issues)
    metric("Issues with an UNRECOGNISED AI label", lambda p: sum(
        1 for r in p.issues.values() if r["attribution"].get("has_ai_label_drift")),
        "Marker drift (e.g. PLANNER_BY_COPILOT, DEV_BY_COPILOT, COPILOT_TESTING). "
        "NOT counted as AI -- each one subtracts from the AI figure until reconciled",
        basis=has_issues)

    ws.append([])
    ws.append(["PULL REQUESTS"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    metric("PRs authored", lambda p: len(p.prs), basis=any_scm)
    metric("PRs merged", lambda p: sum(1 for v in p.prs.values()
                                       if v["event_type"] == "scm.pr.merged"),
           basis=has_prs)
    metric("PRs declined", lambda p: sum(1 for v in p.prs.values()
                                         if v["event_type"] == "scm.pr.declined"),
           basis=has_prs)
    metric("Lines added", lambda p: sum(v["attrs"].get("lines_added") or 0
                                        for v in p.prs.values()), basis=has_prs)
    metric("Lines removed", lambda p: sum(v["attrs"].get("lines_removed") or 0
                                          for v in p.prs.values()), basis=has_prs)
    metric("Files changed", lambda p: sum(v["attrs"].get("files_changed") or 0
                                          for v in p.prs.values()), basis=has_prs)
    metric("Commits", lambda p: sum(v["attrs"].get("commit_count") or 0
                                    for v in p.prs.values()), basis=has_prs)
    metric("Median merge lead time (h)", lambda p: med(
        [ms_to_hours(v["attrs"].get("merge_lead_time_ms")) for v in p.prs.values()]),
        basis=has_prs)
    metric("p85 merge lead time (h)", lambda p: p85(
        [ms_to_hours(v["attrs"].get("merge_lead_time_ms")) for v in p.prs.values()]),
        "The tail is where review actually costs time", basis=has_prs)
    metric("PRs with at least one reviewer", lambda p: sum(
        1 for v in p.prs.values() if (v["attrs"].get("reviewer_count") or 0) > 0),
        "A zero here is a real zero: every pull request was merged unreviewed",
        basis=has_prs)
    metric("Review comments received", lambda p: sum(
        v["attrs"].get("comment_count") or 0 for v in p.prs.values()), basis=has_prs)
    metric("Reviews given to others", lambda p: len(p.reviews), basis=any_scm)
    metric("Reverts touching their commits", lambda p: len(p.reverts),
           "A revert is a signal to read, not a defect count", basis=any_scm)
    metric("PRs with an AI provenance marker", lambda p: sum(
        1 for v in p.prs.values() if v["attrs"].get("has_ai_marker")),
        "AUTH_BY_COPILOT / GEN_BY_COPILOT on a commit subject, an AI-Run-Id trailer, "
        "or [Authored By Copilot] in the PR title", basis=has_prs)

    ws.append([])
    ws.append(["TEST EXECUTION (AIO TCMS)"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    has_tests = lambda p: bool(p.test_runs or p.test_assigned)  # noqa: E731
    metric("Test runs executed", lambda p: test_totals(p)["executed"],
           "Excludes 'Not Run' — AIO seeds every new cycle with those",
           basis=has_tests)
    metric("  passed", lambda p: test_totals(p)["passed"], basis=has_tests)
    metric("  failed", lambda p: test_totals(p)["failed"], basis=has_tests)
    metric("  blocked", lambda p: test_totals(p)["blocked"], basis=has_tests)
    metric("  skipped", lambda p: test_totals(p)["skipped"], basis=has_tests)

    def pass_rate(p: PersonData) -> Optional[float]:
        totals = test_totals(p)
        # Blocked and skipped stay in the denominator: a blocked test is a test
        # that could not be run, and hiding it flatters the number.
        return (round(100 * totals["passed"] / totals["executed"], 1)
                if totals["executed"] else None)

    metric("Pass rate %", pass_rate,
           "passed / executed. Blocked and skipped stay in the denominator",
           basis=has_tests)
    metric("Distinct cycles worked in", lambda p: len(
        {r["attrs"].get("test_cycle_key") for r in p.test_runs
         if r["attrs"].get("test_cycle_key")}), basis=has_tests)
    metric("Automated runs", lambda p: sum(
        1 for r in p.test_runs if r["attrs"].get("is_automated")),
        "Manual vs automated is AIO's own flag, not an inference", basis=has_tests)
    metric("Defects raised from runs", lambda p: sum(
        r["attrs"].get("defect_count") or 0 for r in p.test_runs), basis=has_tests)
    metric("Cases assigned but run by someone else",
           lambda p: len(p.test_assigned),
           "Assignment is not execution and is counted separately",
           basis=has_tests)

    ws.append([])
    ws.append(["AUTOMATION (test case inventory)"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    owns_cases = lambda p: bool(p.owned_cases)  # noqa: E731
    metric("Test cases owned", lambda p: coverage_totals(p)["owned"],
           basis=owns_cases)
    metric("  automated", lambda p: coverage_totals(p)["automated"],
           basis=owns_cases)
    metric("  to be automated", lambda p: coverage_totals(p)["to_be_automated"],
           basis=owns_cases)
    metric("  automation status not set",
           lambda p: coverage_totals(p)["status_unset"],
           "Excluded from the coverage rate — unset is not 'not automated'",
           basis=owns_cases)

    def coverage_pct(p: PersonData) -> Any:
        """Coverage over known-status cases, flagged when the unknowns dominate.

        A bare "100%" is the failure mode to guard against here: it happens when
        every case with a status says Automated while a larger pile has no status
        at all. The percentage is arithmetically right and tells the reader
        nothing, so it is annotated rather than printed clean.
        """
        totals = coverage_totals(p)
        if not totals["known"]:
            return None
        pct = round(100 * totals["automated"] / totals["known"], 1)
        if totals["status_unset"] > totals["known"]:
            return (f"{pct}% (unreliable: {totals['status_unset']} of "
                    f"{totals['owned']} cases have no status)")
        return pct

    metric("Automation coverage %", coverage_pct,
           "automated / (automated + to be automated). Cases with no status set "
           "are excluded, and the figure is flagged when they outnumber the ones "
           "that have one",
           basis=owns_cases)
    metric("Automation scripts created",
           lambda p: sum(v["attrs"].get("automation_scripts_added") or 0
                         for v in p.prs.values()),
           ".feature / step-definition / spec files added in their PRs",
           basis=has_prs)
    metric("Automation scripts modified",
           lambda p: sum(v["attrs"].get("automation_scripts_modified") or 0
                         for v in p.prs.values()), basis=has_prs)

    ws.append([])
    ws.append(["NOT MEASURED IN THIS FILE"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    for gap_row in gaps:
        ws.append([gap_row[0], gap_row[1]])
        ws.cell(row=ws.max_row, column=2).font = NOTE_FONT

    ws.freeze_panes = f"A{header_row + 1}"
    ws.column_dimensions["A"].width = 40
    for idx in range(2, len(names) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    ws.column_dimensions[get_column_letter(len(names) + 2)].width = 62


def build_issue_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    rows = []
    for name, person in data.items():
        for key, rec in sorted(person.issues.items()):
            issue = rec["issue"]
            attribution = rec["attribution"]
            created = parse_ts(issue.get("created_at"))
            resolved = parse_ts(issue.get("resolved_at"))
            cycle = round((resolved - created).total_seconds() / 3600, 1) \
                if created and resolved else None
            rows.append([
                name, key, issue.get("issue_type"), issue.get("status"),
                issue.get("status_category"), issue.get("priority"),
                ", ".join(sorted(rec["roles"])),
                rec["transitions"] or None,
                (issue.get("created_at") or "")[:19].replace("T", " "),
                (issue.get("resolved_at") or "")[:19].replace("T", " ") or "not resolved",
                na(cycle),
                fmt_list(issue.get("labels")),
                "yes" if attribution.get("has_ai_labels") else "no",
                fmt_list(attribution.get("unrecognised_ai_labels")),
                attribution.get("feature_ticket_key") or "",
                issue.get("parent_key") or "",
                round((issue.get("time_spent_seconds") or 0) / 3600, 2)
                if issue.get("time_spent_seconds") else "n/a",
            ])
    write_sheet(
        wb, "Jira Issues",
        ["Person", "Issue", "Type", "Status", "Category", "Priority", "Role(s)",
         "Transitions", "Created", "Resolved", "Cycle time (h)", "Labels",
         "AI label", "Unrecognised AI labels", "Feature ticket (AR-3)", "Parent",
         "Time logged (h)"],
        rows,
        "Cycle time is created -> resolved. 'not resolved' is left as text so an "
        "open issue never averages in as a zero. Time logged is n/a unless someone "
        "actually filled in a worklog -- most rows will be n/a, and that is the "
        "honest answer rather than an implied zero.")


def build_transition_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    rows = []
    for name, person in data.items():
        for t in sorted(person.transitions, key=lambda x: x["transitioned_at"] or ""):
            rows.append([
                name, t["issue_key"],
                t["from_status"] or ("(created)" if t["synthesised"] else ""),
                t["to_status"], t["status_category"],
                (t["transitioned_at"] or "")[:19].replace("T", " "),
                na(t["age_hours"]),
            ])
    write_sheet(
        wb, "Jira Transitions",
        ["Person", "Issue", "From", "To", "Category", "At", "Issue age (h)"],
        rows,
        "Only transitions on issues assigned to the named person are listed. "
        "'(created)' is a synthesised first transition -- Jira's changelog does "
        "not record issue creation, so the poller reconstructs it.")


def build_pr_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    rows = []
    for name, person in data.items():
        for pr_id, rec in sorted(person.prs.items(), key=lambda kv: kv[0] or 0):
            a = rec["attrs"]
            rows.append([
                name, pr_id, rec["ctx"].get("repo_full_name"),
                a.get("pr_state"), rec["ctx"].get("branch_name"),
                rec["ctx"].get("jira_issue_key") or "",
                (a.get("created_on") or "")[:19].replace("T", " "),
                (a.get("merged_at") or a.get("declined_at") or "")[:19].replace("T", " "),
                na(ms_to_hours(a.get("merge_lead_time_ms")
                               or a.get("decline_lead_time_ms"))),
                na(ms_to_hours(a.get("review_lead_time_ms"))),
                na(ms_to_hours(a.get("review_to_merge_ms"))),
                a.get("commit_count"), a.get("files_changed"),
                a.get("lines_added"), a.get("lines_removed"),
                a.get("reviewer_count"), a.get("approval_count"),
                a.get("changes_requested_count"),
                a.get("comment_count"), a.get("inline_comment_count"),
                a.get("commits_after_first_review"),
                "yes" if a.get("has_ai_marker") else "no",
                a.get("ai_commit_count"),
                fmt_list(a.get("ai_model_ids")),
                a.get("decline_reason_class") or "",
            ])
    write_sheet(
        wb, "Pull Requests",
        ["Person", "PR", "Repo", "State", "Branch", "Jira key", "Created",
         "Closed", "Lead time (h)", "Time to 1st review (h)", "Review->merge (h)",
         "Commits", "Files", "+Lines", "-Lines", "Reviewers", "Approvals",
         "Changes requested", "Comments", "Inline comments",
         "Commits after review", "AI marker", "AI commits", "AI models",
         "Decline reason"],
        rows,
        "PR titles are deliberately absent -- the event schema carries only "
        "whether the title held an AI marker, never its text (design §11.3). "
        "A lead time near zero on a PR with zero reviewers is a self-merge, not "
        "a fast review.")


def build_review_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    rows = []
    for name, person in data.items():
        for rec in person.reviews:
            a = rec["attrs"]
            rows.append([
                name, a.get("pr_id"), rec["ctx"].get("repo_full_name"),
                a.get("action"),
                "yes" if a.get("is_first_review") else "no",
                (a.get("reviewed_at") or "")[:19].replace("T", " "),
                na(ms_to_hours(a.get("review_lead_time_ms"))),
                a.get("comment_count"),
            ])
    write_sheet(
        wb, "Reviews Given",
        ["Person", "PR", "Repo", "Action", "First review", "At",
         "Time from PR open (h)", "Comments"],
        rows,
        "Reviews the named person gave on other people's pull requests. "
        "Empty means Bitbucket recorded no review activity from them in the "
        "window -- it does not mean they reviewed nothing offline.")


def build_revert_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    rows = []
    for name, person in data.items():
        for rec in person.reverts:
            a = rec["attrs"]
            rows.append([
                name, rec["ctx"].get("repo_full_name"),
                (a.get("revert_commit_sha") or "")[:12],
                (a.get("reverted_commit_sha") or "")[:12],
                (a.get("reverted_at") or "")[:19].replace("T", " "),
                na(a.get("days_to_revert")),
                "yes" if a.get("reverted_commit_has_ai_marker") else "no",
                a.get("resolution"),
            ])
    write_sheet(
        wb, "Reverts",
        ["Person", "Repo", "Revert commit", "Reverted commit", "At",
         "Days to revert", "Reverted commit was AI-marked", "Resolution"],
        rows,
        "A revert is not automatically a defect: reverting a merge to re-land it "
        "cleanly is normal. Read the resolution column before drawing a "
        "conclusion, and never turn this count into a quality score.")


def build_test_run_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    rows = []
    for name, person in data.items():
        for rec in sorted(person.test_runs,
                          key=lambda r: r["attrs"].get("executed_at") or ""):
            a = rec["attrs"]
            rows.append([
                name, a.get("test_cycle_key"), a.get("test_case_key"),
                a.get("status"), a.get("status_category"),
                "automated" if a.get("is_automated") else "manual",
                (a.get("executed_at") or "")[:19].replace("T", " ") or "never run",
                a.get("folder_name") or "", a.get("priority") or "",
                a.get("defect_count"),
                round(a["effort_seconds"] / 60, 1) if a.get("effort_seconds")
                else "n/a",
            ])
    write_sheet(
        wb, "Test Runs",
        ["Person", "Cycle", "Test case", "Status", "Category", "Mode",
         "Executed at", "Folder", "Priority", "Defects", "Effort (min)"],
        rows,
        "One row per test case per cycle, for the person who executed it. Test "
        "case titles are not carried — they are free text and often quote "
        "customer names and endpoints (design §11.3); the folder is a "
        "controlled taxonomy and is safe. Effort is n/a unless someone recorded "
        "it, which is almost never; it is not zero.")


def build_cycle_sheet(wb: Workbook, data: Dict[str, PersonData]) -> None:
    """Per-cycle rollup — the view a QA engineer actually recognises as their week."""
    rows = []
    for name, person in data.items():
        by_cycle: Dict[str, Counter] = defaultdict(Counter)
        window: Dict[str, List[str]] = defaultdict(list)
        for rec in person.test_runs:
            a = rec["attrs"]
            key = a.get("test_cycle_key") or "(none)"
            by_cycle[key][a.get("status_category") or "other"] += 1
            by_cycle[key]["total"] += 1
            if a.get("is_automated"):
                by_cycle[key]["automated"] += 1
            by_cycle[key]["defects"] += a.get("defect_count") or 0
            if a.get("executed_at"):
                window[key].append(a["executed_at"])
        for key, counts in sorted(by_cycle.items()):
            executed = sum(counts[c] for c in TEST_EXECUTED)
            stamps = sorted(window[key])
            rows.append([
                name, key, counts["total"], executed,
                counts["passed"], counts["failed"], counts["blocked"],
                counts["skipped"],
                round(100 * counts["passed"] / executed, 1) if executed else "n/a",
                counts["automated"], counts["defects"],
                (stamps[0] or "")[:19].replace("T", " ") if stamps else "n/a",
                (stamps[-1] or "")[:19].replace("T", " ") if stamps else "n/a",
            ])
    write_sheet(
        wb, "Test Cycles",
        ["Person", "Cycle", "Rows", "Executed", "Passed", "Failed", "Blocked",
         "Skipped", "Pass rate %", "Automated", "Defects", "First run", "Last run"],
        rows,
        "'Rows' is every row this person touched in the cycle; 'Executed' is the "
        "subset that actually ran and is the only honest pass-rate denominator. "
        "A cycle spanning weeks between first and last run is normal for "
        "regression work and is not a sign of a stalled cycle.")


def iso_week(dt: Optional[datetime]) -> Optional[str]:
    if dt is None:
        return None
    year, week, _ = dt.isocalendar()
    return f"{year}-W{week:02d}"


def week_start(label: str) -> str:
    year, week = int(label[:4]), int(label[6:])
    return datetime.fromisocalendar(year, week, 1).strftime("%Y-%m-%d")


def week_range(labels: Iterable[str]) -> List[str]:
    """Every ISO week between the first and last observed, gaps included.

    A week with no activity has to appear as a zero row, not be skipped. Dropping
    it would join the week before to the week after and draw a flat line across a
    holiday or an outage, which is the one thing a trend chart must not do.
    """
    labels = sorted(set(labels))
    if not labels:
        return []
    first = datetime.strptime(week_start(labels[0]), "%Y-%m-%d")
    last = datetime.strptime(week_start(labels[-1]), "%Y-%m-%d")
    out, cursor = [], first
    while cursor <= last:
        out.append(iso_week(cursor))
        cursor += timedelta(days=7)
    return out


#: (column label, note). Order drives both the table and the charts.
TREND_METRICS = [
    "Issues resolved", "Issues raised", "PRs merged",
    "Automation scripts created", "Lines changed",
    "Median merge lead time (h)", "PRs reviewed by someone",
    "Tests executed", "Pass rate %", "Tests automated %", "Defects raised",
]


def weekly_trend(person: PersonData,
                 since: Optional[datetime] = None) -> Dict[str, Dict[str, Any]]:
    """Per-ISO-week metrics for one person, gaps filled with zeros.

    ``since`` clamps which *facts* are bucketed. It matters because a fact's
    trend date is not its event date: an issue raised in April can still be
    transitioned in August, so its ``created_at`` would otherwise stretch the
    chart back over ten empty weeks and squash the part anyone wants to read.
    """
    floor = iso_week(since) if since else None

    def keep(label: Optional[str]) -> Optional[str]:
        if label is None:
            return None
        return label if (floor is None or label >= floor) else None
    weeks: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
        "lead_times": [], "issues_resolved": 0, "issues_raised": 0,
        "prs_merged": 0, "lines": 0, "prs_reviewed": 0, "scripts": 0,
        "passed": 0, "executed": 0, "automated": 0, "defects": 0,
    })

    for rec in person.issues.values():
        issue = rec["issue"]
        if "assignee" in rec["roles"]:
            label = keep(iso_week(parse_ts(issue.get("resolved_at"))))
            if label:
                weeks[label]["issues_resolved"] += 1
        if "reporter" in rec["roles"]:
            label = keep(iso_week(parse_ts(issue.get("created_at"))))
            if label:
                weeks[label]["issues_raised"] += 1

    for rec in person.prs.values():
        a = rec["attrs"]
        label = keep(iso_week(parse_ts(a.get("merged_at") or a.get("declined_at")
                                       or a.get("created_on"))))
        if not label:
            continue
        bucket = weeks[label]
        if rec["event_type"] == "scm.pr.merged":
            bucket["prs_merged"] += 1
        bucket["lines"] += (a.get("lines_added") or 0) + (a.get("lines_removed") or 0)
        bucket["scripts"] += a.get("automation_scripts_added") or 0
        if (a.get("reviewer_count") or 0) > 0:
            bucket["prs_reviewed"] += 1
        lead = ms_to_hours(a.get("merge_lead_time_ms"))
        if lead is not None:
            bucket["lead_times"].append(lead)

    for rec in person.test_runs:
        a = rec["attrs"]
        label = keep(iso_week(parse_ts(a.get("executed_at"))))
        if not label:
            continue
        bucket = weeks[label]
        category = a.get("status_category")
        if category in TEST_EXECUTED:
            bucket["executed"] += 1
            if category == "passed":
                bucket["passed"] += 1
            if a.get("is_automated"):
                bucket["automated"] += 1
        bucket["defects"] += a.get("defect_count") or 0

    out: Dict[str, Dict[str, Any]] = {}
    for label in week_range(weeks):
        b = weeks.get(label) or {
            "lead_times": [], "issues_resolved": 0, "issues_raised": 0,
            "prs_merged": 0, "lines": 0, "prs_reviewed": 0, "scripts": 0,
            "passed": 0, "executed": 0, "automated": 0, "defects": 0,
        }
        executed = b["executed"]
        out[label] = {
            "Issues resolved": b["issues_resolved"],
            "Issues raised": b["issues_raised"],
            "PRs merged": b["prs_merged"],
            "Automation scripts created": b["scripts"],
            "Lines changed": b["lines"],
            # None, not 0: a week with no merged PR has no lead time to report,
            # and a 0 would draw the line down to the axis as though every PR
            # merged instantly.
            "Median merge lead time (h)": med(b["lead_times"]),
            "PRs reviewed by someone": b["prs_reviewed"],
            "Tests executed": executed,
            "Pass rate %": (round(100 * b["passed"] / executed, 1)
                            if executed else None),
            "Tests automated %": (round(100 * b["automated"] / executed, 1)
                                  if executed else None),
            "Defects raised": b["defects"],
        }
    return out


def build_trend_sheet(wb: Workbook, data: Dict[str, PersonData],
                      since: Optional[datetime] = None) -> Dict[str, Any]:
    """Long-format weekly table. Returns the trend data for the chart sheet."""
    trends = {name: weekly_trend(person, since) for name, person in data.items()}
    all_weeks = week_range(w for t in trends.values() for w in t)

    # The week in progress is always short, so it always looks like a decline.
    # It is kept -- a manager wants to see the current week -- but labelled, and
    # excluded from the charts so a half-week never reads as a downward trend.
    current = iso_week(datetime.now(timezone.utc))
    complete_weeks = [w for w in all_weeks if w != current]

    rows = []
    for name in data:
        for label in all_weeks:
            metrics = trends[name].get(label)
            if metrics is None:
                continue
            rows.append([name, label + (" (partial)" if label == current else ""),
                         week_start(label)] +
                        [na(metrics[m]) for m in TREND_METRICS])
    write_sheet(
        wb, "Weekly Trend", ["Person", "Week", "Week starting"] + TREND_METRICS, rows,
        "One row per person per ISO week, including weeks with no activity — a "
        "skipped week would join the week before to the week after and draw a "
        "flat line across a holiday. Rates and lead times are 'n/a' when the "
        "week had nothing to divide by; they are not zero. The week marked "
        "(partial) is still in progress and is excluded from the charts, because "
        "a half-finished week always looks like a decline. Four or five weeks is "
        "context, not a trend: read the slope over eight or more.")
    return {"trends": trends, "weeks": complete_weeks, "partial": current}


def build_chart_sheet(wb: Workbook, data: Dict[str, PersonData],
                      trend: Dict[str, Any]) -> None:
    """One small multiple per metric, a series per person -- the management view.

    The series data lives on a separate **visible** sheet rather than in hidden
    columns beside the charts. That is not a style choice: Excel writes
    ``plotVisOnly=1`` by default, so a chart whose source columns are hidden
    renders as an empty frame. The first version of this function hid them and
    produced ten blank charts.
    """
    trends, weeks = trend["trends"], trend["weeks"]
    if not weeks:
        return

    names = list(data)
    src = wb.create_sheet("Trend Data")
    src.append(["Series data for the Trend Charts sheet. Editing it changes the "
                "charts; the Weekly Trend sheet is the one to read."])
    src.cell(row=1, column=1).font = NOTE_FONT

    ws = wb.create_sheet("Trend Charts", 1)
    ws.append(["Weekly trend"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"One chart per metric over ISO weeks, one line per person. The "
               f"week in progress ({trend.get('partial')}) is excluded — a "
               f"partial week always looks like a decline; see the Weekly Trend "
               f"sheet for it. A gap in a line is a week with nothing to divide "
               f"by, not a zero."])
    ws.cell(row=2, column=1).font = NOTE_FONT

    row_cursor = 3
    chart_row = 4
    for index, metric in enumerate(TREND_METRICS):
        header_row = row_cursor
        src.cell(row=header_row, column=1, value=metric).font = SECTION_FONT
        header_row += 1
        src.cell(row=header_row, column=1, value="Week")
        for offset, name in enumerate(names, start=1):
            src.cell(row=header_row, column=1 + offset, value=name)
        for week_index, label in enumerate(weeks, start=1):
            src.cell(row=header_row + week_index, column=1, value=label)
            for offset, name in enumerate(names, start=1):
                src.cell(row=header_row + week_index, column=1 + offset,
                         value=(trends[name].get(label) or {}).get(metric))

        chart = LineChart()
        chart.title = metric
        chart.height, chart.width = 7.5, 15
        chart.y_axis.title = metric
        chart.x_axis.title = "ISO week"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        # Belt and braces alongside the visible source sheet: never let a hidden
        # row or column silently blank a chart again. (openpyxl spells the
        # plotVisOnly flag `visible_cells_only`; setting `plotVisOnly` directly
        # is silently ignored and leaves it at the default of 1.)
        chart.visible_cells_only = False
        # A missing week is a gap in the line, not a drop to zero.
        chart.display_blanks = "gap"
        chart.add_data(
            Reference(src, min_col=2, max_col=1 + len(names),
                      min_row=header_row, max_row=header_row + len(weeks)),
            titles_from_data=True)
        # Categories are ISO-week labels -- text. set_categories() would write
        # them as a numRef, which Excel is entitled to render as 1, 2, 3 instead
        # of 2026-W28, so the reference is forced to a string reference.
        categories = AxDataSource(strRef=StrRef(
            f"'{src.title}'!$A${header_row + 1}:$A${header_row + len(weeks)}"))
        for series in chart.series:
            series.cat = categories
            series.smooth = False   # straight segments between real points
        ws.add_chart(chart, f"{get_column_letter(1 + (index % 2) * 9)}{chart_row}")
        if index % 2 == 1:
            chart_row += 16
        row_cursor = header_row + len(weeks) + 2

    src.column_dimensions["A"].width = 28
    for offset in range(1, len(names) + 1):
        src.column_dimensions[get_column_letter(1 + offset)].width = 18
    ws.column_dimensions["A"].width = 14


def build_coverage_sheet(wb: Workbook, sources: Sequence[Sequence[Any]],
                         gaps: Sequence[Sequence[str]],
                         stats: Dict[str, int]) -> None:
    ws = wb.create_sheet("Coverage & Gaps")
    ws.append(["What this file is built from"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Source", "Status", "Events", "Detail"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in sources:
        ws.append(list(row))

    ws.append([])
    ws.append(["What is NOT in this file, and why"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=13)
    ws.append(["Missing", "Reason"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in gaps:
        ws.append(list(row))

    ws.append([])
    ws.append(["Data quality"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=13)
    ws.append(["Files read", stats["files"]])
    ws.append(["Lines read", stats["lines"]])
    ws.append(["Malformed lines dropped", stats["malformed"]])
    ws.append(["Duplicate event ids dropped", stats["duplicates"]])

    ws.append([])
    ws.append(["How to read this workbook"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=13)
    for line in (
        "An empty sheet means no source events, not a measurement of zero.",
        "Lead times are medians and p85, never means -- these distributions have "
        "long tails and a mean hides them.",
        "Volume columns (lines, commits, PR counts) describe throughput. They are "
        "not productivity and not quality; a large diff is often a generated "
        "fixture, and a small one is often the hard bug.",
        "There is no AI-vs-no-AI comparison here. Essentially all work in this "
        "org now uses AI, so there is no control group and any such number would "
        "be a model presented as a measurement (design §8.16).",
        "Two people is below any threshold at which a percentage means anything. "
        "Read the rows, not the ratios.",
    ):
        ws.append([line])
        ws.cell(row=ws.max_row, column=1).font = NOTE_FONT

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Per-person Excel workbook from AI telemetry NDJSON.")
    ap.add_argument("--person", action="append", required=True,
                    metavar="NAME=ACCOUNT_ID",
                    help="Repeatable. Only named people appear in the file.")
    ap.add_argument("--input", nargs="+", required=True,
                    help="NDJSON files or directories.")
    ap.add_argument("--since", help="YYYY-MM-DD window start (inclusive).")
    ap.add_argument("--out", required=True, help="Output .xlsx path.")
    ap.add_argument("--source", action="append", default=[],
                    metavar="NAME|STATUS|DETAIL",
                    help="Repeatable Coverage sheet row.")
    ap.add_argument("--gap", action="append", default=[],
                    metavar="MISSING|REASON",
                    help="Repeatable 'not measured' row.")
    args = ap.parse_args(argv)

    people: Dict[str, str] = {}
    for spec in args.person:
        if "=" not in spec:
            return _fail(f"--person needs NAME=ACCOUNT_ID, got: {spec}")
        name, _, aid = spec.partition("=")
        people[name.strip()] = aid.strip()

    since = None
    if args.since:
        since = parse_ts(args.since + "T00:00:00Z")
        if since is None:
            return _fail(f"--since is not a date: {args.since}")

    events, stats = load_events(args.input)
    if not events:
        return _fail("no events found in the given input")
    data = collect(events, people, since)

    counts = Counter(e["event_type"] for e in events)
    sources = [row.split("|") for row in args.source] or [
        [k, "collected", v, ""] for k, v in sorted(counts.items())
    ]
    gaps = [row.split("|") for row in args.gap]

    window = f"window: {args.since or 'all data'} .. today"
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, data, window, gaps)
    build_issue_sheet(wb, data)
    build_transition_sheet(wb, data)
    build_pr_sheet(wb, data)
    build_review_sheet(wb, data)
    build_revert_sheet(wb, data)
    build_cycle_sheet(wb, data)
    build_test_run_sheet(wb, data)
    trend = build_trend_sheet(wb, data, since)
    build_chart_sheet(wb, data, trend)
    build_coverage_sheet(wb, sources, gaps, stats)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)) or ".", exist_ok=True)
    wb.save(args.out)

    print(json.dumps({
        "msg": "workbook_written", "out": args.out,
        "people": {n: {"issues": len(p.issues), "prs": len(p.prs),
                       "reviews": len(p.reviews), "reverts": len(p.reverts),
                       "test_runs": len(p.test_runs),
                       "test_executed": test_totals(p)["executed"]}
                   for n, p in data.items()},
        "events_read": len(events), **stats,
    }, sort_keys=True))
    return 0


def _fail(message: str) -> int:
    print(f"error: {message}", file=sys.stderr)
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
