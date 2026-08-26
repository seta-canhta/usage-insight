"""Tests for combined_weekly.py -- stdlib unittest, no network, no fixtures."""

from __future__ import annotations

import contextlib
import importlib.util
import io
import json
import os
import shutil
import sys
import tempfile
import unittest
from datetime import datetime, timedelta, timezone

_HERE = os.path.dirname(os.path.abspath(__file__))
_SPEC = importlib.util.spec_from_file_location(
    "combined_weekly", os.path.join(os.path.dirname(_HERE), "combined_weekly.py")
)
cw = importlib.util.module_from_spec(_SPEC)
sys.modules["combined_weekly"] = cw

# openpyxl is an optional dependency: `people_workbook.py` guards its import and
# exits when it is missing, so importing this module on a machine without it
# aborts the whole `report` suite rather than reporting one absent tool. A tool
# that is not installed is a skip, not a failure -- and rendering it as a
# failure is the same mistake this project refuses to make with data.
#
#     python3 -m pip install openpyxl     # to actually run them
try:
    _SPEC.loader.exec_module(cw)
    from openpyxl import load_workbook  # noqa: E402
    HAS_OPENPYXL = True
except (ImportError, SystemExit):
    HAS_OPENPYXL = False
    load_workbook = None


def load_tests(loader, tests, pattern):     # noqa: D103 -- unittest protocol
    return tests if HAS_OPENPYXL else unittest.TestSuite()


ALPHA = "0123456789abcdef01234567"
BETA = "712020:00000000-1111-2222-3333-444444444444"


def base(event_type, actor, when, attrs, event_id=None):
    return {
        "schema_version": "1.0.0",
        "event_id": event_id or f"evt_{event_type}_{actor}_{when}",
        "event_type": event_type,
        "event_time": when,
        "actor": {"person_id": actor, "person_email_hash": None},
        "context": {"repo_full_name": "ws/repo", "branch_name": "b",
                    "jira_issue_key": None, "jira_project_key": "PROJ"},
        "agent": {"agent_name": "poller.test", "surface": "headless"},
        "attributes": attrs,
        "link": {"method": "heuristic", "confidence": 0.5},
    }


def pr(actor, pr_id, when, reviewers=0, scripts=0, merged=True):
    kind = "scm.pr.merged" if merged else "scm.pr.declined"
    attrs = {"pr_id": pr_id, "pr_state": "MERGED" if merged else "DECLINED",
             "merged_at": when if merged else None,
             "declined_at": None if merged else when,
             "created_on": when, "merge_lead_time_ms": 3_600_000,
             "lines_added": 10, "lines_removed": 2, "files_changed": 3,
             "commit_count": 1, "reviewer_count": reviewers, "comment_count": 0,
             "automation_scripts_added": scripts,
             "automation_scripts_modified": 1,
             "automation_files_by_kind": {"feature": {"added": scripts}},
             "has_ai_marker": False}
    return base(kind, actor, when, attrs, f"evt_pr_{pr_id}")


def run(actor, case, when, cycle="CY-1", category="passed"):
    return base("test.run.completed", actor, when, {
        "test_case_key": case, "test_cycle_key": cycle, "test_run_id": 1,
        "status": category.title(), "status_category": category,
        "is_automated": True, "executed_by_person_id": actor,
        "assigned_to_person_id": actor, "executed_at": when,
        "effort_seconds": None, "defect_count": 1,
        "folder_name": "Regression", "priority": "High",
    }, f"evt_run_{cycle}_{case}")


def snapshot(actor, case, status, when, priority="High"):
    return base("test.case.snapshot", actor, when, {
        "test_case_key": case, "automation_status": status,
        "automation_owner_person_id": None, "has_automation_key": False,
        "test_case_status": "Published", "script_type": "Classic",
        "folder_name": "Regression", "priority": priority,
        "is_archived": False, "created_at": when, "updated_at": when,
    }, f"evt_case_{case}")


def jira_transition(actor, key, when, attribution=None):
    issue = {"issue_key": key, "issue_type": "Bug", "status": "Done",
             "assignee_person_id": actor, "reporter_person_id": actor,
             "created_at": when, "resolved_at": None, "labels": [],
             "time_spent_seconds": None}
    return base("jira.transition", actor, when, {
        "issue": issue, "jira_issue_key": key, "to_status": "Done",
        "from_status": "In Progress", "status_category": "done",
        "transitioned_at": when, "age_at_transition_ms": 7_200_000,
        "attribution": attribution or {"has_ai_labels": False},
    }, f"evt_tr_{key}_{actor}")


class TestRenderHelpers(unittest.TestCase):
    def test_none_renders_as_no_data_not_zero(self):
        self.assertEqual(cw.cell(None), "no data")
        self.assertEqual(cw.cell(0), "0")

    def test_large_integers_are_grouped(self):
        self.assertEqual(cw.cell(261678), "261,678")

    def test_floats_lose_trailing_zeros(self):
        self.assertEqual(cw.cell(1.50), "1.5")
        self.assertEqual(cw.cell(2.0), "2")

    def test_empty_table_renders_nothing(self):
        self.assertEqual(cw.table(["A", "B"], []), [])

    def test_pct_keeps_absent_distinct(self):
        self.assertEqual(cw.pct(None), "no data")
        self.assertEqual(cw.pct(97.6), "97.6%")


class TestReport(unittest.TestCase):
    def _events(self):
        # Two ISO weeks so the trend has something to show.
        w32, w33 = "2026-08-05T10:00:00Z", "2026-08-12T10:00:00Z"
        return [
            pr(ALPHA, 1, w32, reviewers=1, scripts=2),
            pr(BETA, 2, w32, reviewers=0, scripts=0),
            pr(BETA, 3, w33, reviewers=0, scripts=0),
            run(ALPHA, "TC-1", w33),
            run(BETA, "TC-2", w33, category="failed"),
            snapshot(ALPHA, "TC-1", "Automated", w33),
            snapshot(ALPHA, "TC-3", "To Be Automated", w33),
        ]

    def _render(self, events=None, extra_argv=()):
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, "e.ndjson")
        with open(src, "w") as fh:
            for e in (events if events is not None else self._events()):
                fh.write(json.dumps(e) + "\n")
        out = os.path.join(tmp, "report.md")
        with contextlib.redirect_stdout(io.StringIO()):
            rc = cw.main(["--person", f"Alpha={ALPHA}", "--person", f"Beta={BETA}",
                          "--input", src, "--week", "2026-W33",
                          "--since", "2026-08-01", "--out", out, *extra_argv])
        self.assertEqual(rc, 0)
        return open(out, encoding="utf-8").read()

    def test_one_file_holds_people_and_project(self):
        body = self._render()
        self.assertIn("# Weekly report", body)
        self.assertIn("## 1. At a glance", body)
        self.assertIn("## 2. Weekly trend", body)
        self.assertIn("## 3. Alpha", body)
        self.assertIn("## 4. Beta", body)
        self.assertIn("Project — business as usual", body)

    def test_week_and_window_totals_are_separated(self):
        # Window totals under a heading naming one week is how a reader comes to
        # believe 53 pull requests were merged in five days.
        body = self._render()
        self.assertIn("**This week — 2026-W33**", body)
        self.assertIn("**Whole window — 2026-08-01 → today**", body)

    def test_unreviewed_merges_are_called_out(self):
        body = self._render()
        self.assertIn("none of them reviewed", body)
        self.assertIn("no first review to measure", body)

    def test_unrecognised_ai_labels_are_reported_next_to_the_ai_count(self):
        """Drift never folds into the AI figure -- it is the gap in that figure."""
        events = self._events() + [
            jira_transition(ALPHA, "PROJ-1", "2026-08-12T10:00:00Z", attribution={
                "has_ai_labels": True,
                "unrecognised_ai_labels": ["DEV_BY_COPILOT"],
                "has_ai_label_drift": True,
            }),
            jira_transition(BETA, "PROJ-2", "2026-08-12T10:00:00Z"),
        ]
        body = self._render(events)
        self.assertIn("Carrying an UNRECOGNISED AI label", body)
        self.assertIn("| Carrying an AI label | 1 |", body)
        self.assertIn("| Carrying an UNRECOGNISED AI label | 1 |", body)
        self.assertIn("| Carrying an UNRECOGNISED AI label | 0 |", body)

    def test_forbidden_metrics_are_never_rendered_as_values(self):
        body = self._render().lower()
        self.assertNotIn("| roi ", body)
        self.assertNotIn("time saved |", body)
        self.assertNotIn("productivity gain |", body)
        # They may only appear in the section that explains their absence.
        self.assertIn("no control group", body)

    def test_release_section_is_added_and_labelled_incomparable(self):
        tmp = tempfile.mkdtemp()
        rel = os.path.join(tmp, "rel.ndjson")
        with open(rel, "w") as fh:
            fh.write(json.dumps(
                snapshot(ALPHA, "TC-9", "To Be Automated",
                         "2026-08-12T10:00:00Z", priority="Low")) + "\n")
        body = self._render(extra_argv=("--release-input", rel,
                                        "--release-label", "26.8"))
        self.assertIn("Release 26.8", body)
        self.assertIn("not comparable", body)

    def test_scope_note_is_printed(self):
        body = self._render(extra_argv=("--scope-note", "PRJ only"))
        self.assertIn("Scope: PRJ only", body)

    def test_priority_mapping_is_stated_not_assumed(self):
        body = self._render()
        self.assertIn("AIO has no P1/P2/P3 field", body)

    def test_bad_person_spec_is_rejected(self):
        with contextlib.redirect_stderr(io.StringIO()):
            self.assertEqual(cw.main(["--person", "NoEquals",
                                      "--input", os.devnull, "--out", "x.md"]), 2)

    def test_empty_input_is_rejected_rather_than_rendering_zeros(self):
        with tempfile.TemporaryDirectory() as tmp:
            empty = os.path.join(tmp, "empty.ndjson")
            open(empty, "w").close()
            with contextlib.redirect_stderr(io.StringIO()):
                rc = cw.main(["--person", f"Alpha={ALPHA}", "--input", empty,
                              "--out", os.path.join(tmp, "o.md")])
        self.assertEqual(rc, 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)


class TestCoverageSection(unittest.TestCase):
    """Coverage is what lets a reader tell a measured zero from an absent week,
    so its wording is load-bearing rather than decorative."""

    def setUp(self):
        self.dir = tempfile.mkdtemp(prefix="cov-")
        self.addCleanup(shutil.rmtree, self.dir, True)

    def write(self, report):
        path = os.path.join(self.dir, "coverage.json")
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(report, handle)
        return path

    def test_no_coverage_file_says_so_rather_than_reporting_zero(self):
        body = "\n".join(cw.section_coverage(6, None))
        self.assertIn("has not started", body)
        self.assertNotIn("0 machine-week", body)

    def test_a_missing_file_is_treated_the_same_as_none(self):
        body = "\n".join(cw.section_coverage(6, os.path.join(self.dir, "nope.json")))
        self.assertIn("has not started", body)

    def test_one_week_reads_as_singular(self):
        body = "\n".join(cw.section_coverage(6, self.write({
            "machines": 1, "machine_weeks_covered": 1, "weeks_seen": ["2026-W35"],
            "by_machine": {"abc12345def": {
                "weeks_covered": ["2026-W35"], "weeks_missing_within_span": [],
                "bundles": 1, "events": 21}}})))
        self.assertIn("**1 machine-week** across 1 machine,", body)
        self.assertNotIn("machine-weeks", body)
        self.assertNotIn("machine(s)", body)
        self.assertNotIn("2026-W35 to 2026-W35", body)

    def test_missing_weeks_are_named_not_counted(self):
        # "3 weeks missing" invites a shrug. The weeks themselves invite a
        # question about what happened in them.
        body = "\n".join(cw.section_coverage(6, self.write({
            "machines": 1, "machine_weeks_covered": 2,
            "weeks_seen": ["2026-W32", "2026-W35"],
            "by_machine": {"abc12345def": {
                "weeks_covered": ["2026-W32", "2026-W35"],
                "weeks_missing_within_span": ["2026-W33", "2026-W34"],
                "bundles": 2, "events": 40}}})))
        self.assertIn("2026-W33, 2026-W34", body)
        self.assertIn("2026-W32 to 2026-W35", body)
