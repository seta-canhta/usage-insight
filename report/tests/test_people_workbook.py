"""Tests for people_workbook.py -- stdlib unittest plus openpyxl, no network."""

from __future__ import annotations

import contextlib
import importlib.util
import io
import json
import os
import sys
import tempfile
import unittest
from datetime import datetime, timezone

_HERE = os.path.dirname(os.path.abspath(__file__))
_SPEC = importlib.util.spec_from_file_location(
    "people_workbook", os.path.join(os.path.dirname(_HERE), "people_workbook.py")
)
pw = importlib.util.module_from_spec(_SPEC)
sys.modules["people_workbook"] = pw
_SPEC.loader.exec_module(pw)

from openpyxl import load_workbook  # noqa: E402

# Synthetic account ids in both Atlassian shapes -- the legacy 24-hex form and
# the newer "712020:<uuid>" form -- because the code slices and compares them.
# No real person appears in this file.
ALPHA = "0123456789abcdef01234567"
BETA = "712020:00000000-1111-2222-3333-444444444444"
PEOPLE = {"Alpha": ALPHA, "Beta": BETA}


def base(event_type, actor, when="2026-08-12T10:00:00Z", attrs=None, ctx=None,
         event_id=None):
    return {
        "schema_version": "1.0.0",
        "event_id": event_id or f"evt_{event_type}_{actor}_{when}",
        "event_type": event_type,
        "event_time": when,
        "actor": {"person_id": actor, "person_email_hash": None},
        "context": ctx or {"repo_full_name": "ws/repo", "branch_name": "b",
                           "jira_issue_key": None, "jira_project_key": None},
        "agent": {"agent_name": "poller.test", "surface": "headless"},
        "attributes": attrs or {},
        "link": {"method": "heuristic", "confidence": 0.5},
    }


def pr(actor, pr_id, state="scm.pr.merged", **attrs):
    payload = {"pr_id": pr_id, "pr_state": state.split(".")[-1].upper(),
               "lines_added": 10, "lines_removed": 1, "files_changed": 2,
               "commit_count": 1, "reviewer_count": 0, "comment_count": 0,
               "merge_lead_time_ms": 3_600_000, "has_ai_marker": False}
    payload.update(attrs)
    return base(state, actor, attrs=payload, event_id=f"evt_pr_{pr_id}_{state}")


def transition(actor, key, assignee=None, reporter=None, **issue_extra):
    issue = {"issue_key": key, "issue_type": "Bug", "status": "Done",
             "assignee_person_id": assignee, "reporter_person_id": reporter,
             "created_at": "2026-08-10T10:00:00Z", "resolved_at": None,
             "labels": [], "time_spent_seconds": None}
    issue.update(issue_extra)
    return base("jira.transition", actor, attrs={
        "issue": issue, "jira_issue_key": key, "to_status": "Done",
        "from_status": "In Progress", "status_category": "done",
        "transitioned_at": "2026-08-12T10:00:00Z",
        "age_at_transition_ms": 7_200_000,
        "attribution": {"has_ai_labels": False},
    }, event_id=f"evt_tr_{key}_{actor}")


class TestHelpers(unittest.TestCase):
    def test_median_and_p85_return_none_on_empty(self):
        self.assertIsNone(pw.med([]))
        self.assertIsNone(pw.p85([]))
        self.assertEqual(pw.med([1, 2, 3]), 2)
        # Nearest-rank, not interpolated: round(0.85 * 4) == 3 -> the 4th value.
        self.assertEqual(pw.p85([1, 2, 3, 4, 5]), 4)
        self.assertEqual(pw.p85([1]), 1)

    def test_none_is_skipped_not_treated_as_zero(self):
        # A PR with no review has review_lead_time_ms = None. Counting it as 0
        # would report instant reviews that never happened.
        self.assertEqual(pw.med([None, 4, None, 6]), 5)

    def test_ms_to_hours(self):
        self.assertIsNone(pw.ms_to_hours(None))
        self.assertEqual(pw.ms_to_hours(3_600_000), 1.0)

    def test_parse_ts_handles_z_and_naive(self):
        self.assertIsNotNone(pw.parse_ts("2026-08-12T10:00:00Z"))
        self.assertIsNone(pw.parse_ts("nonsense"))
        self.assertIsNone(pw.parse_ts(None))


class TestCollect(unittest.TestCase):
    def test_since_filters_events_out(self):
        evs = [pr(BETA, 1), base("scm.pr.merged", BETA, when="2026-01-01T00:00:00Z",
                                 attrs={"pr_id": 2}, event_id="old")]
        since = datetime(2026, 6, 1, tzinfo=timezone.utc)
        data = pw.collect(evs, PEOPLE, since)
        self.assertEqual(list(data["Beta"].prs), [1])

    def test_roles_are_kept_apart(self):
        # Transitioning someone else's ticket is not the same as owning it.
        evs = [transition(ALPHA, "PROJ-1", assignee=BETA, reporter=ALPHA)]
        data = pw.collect(evs, PEOPLE, None)
        self.assertEqual(data["Beta"].issues["PROJ-1"]["roles"], {"assignee"})
        self.assertEqual(data["Alpha"].issues["PROJ-1"]["roles"],
                         {"reporter", "transitioned"})
        # Only the assignee's timeline gets the transition row.
        self.assertEqual(len(data["Beta"].transitions), 1)
        self.assertEqual(len(data["Alpha"].transitions), 0)

    def test_reviews_are_attributed_to_the_reviewer_not_the_pr_author(self):
        evs = [base("scm.pr.reviewed", ALPHA, attrs={
            "pr_id": 7, "reviewer_person_id": BETA, "action": "approve",
            "comment_count": 1, "is_first_review": True,
            "review_lead_time_ms": 60_000})]
        data = pw.collect(evs, PEOPLE, None)
        self.assertEqual(len(data["Beta"].reviews), 1)
        self.assertEqual(len(data["Alpha"].reviews), 0)

    def test_people_not_on_the_allowlist_are_ignored(self):
        data = pw.collect([pr("someone-else", 9)], PEOPLE, None)
        self.assertEqual(sum(len(p.prs) for p in data.values()), 0)


class TestLoading(unittest.TestCase):
    def test_malformed_and_duplicate_lines_are_accounted_for(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "a.ndjson")
            with open(path, "w") as fh:
                fh.write(json.dumps(pr(BETA, 1)) + "\n")
                fh.write(json.dumps(pr(BETA, 1)) + "\n")   # same event_id
                fh.write("{broken\n")
                fh.write(json.dumps({"no_type": 1}) + "\n")
            events, stats = pw.load_events([path])
        self.assertEqual(len(events), 1)
        self.assertEqual(stats["duplicates"], 1)
        self.assertEqual(stats["malformed"], 2)


class TestWorkbook(unittest.TestCase):
    """The rule this file exists to protect: 0 and 'no data' are different."""

    def _build(self, events):
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, "e.ndjson")
        with open(src, "w") as fh:
            for e in events:
                fh.write(json.dumps(e) + "\n")
        out = os.path.join(tmp, "out.xlsx")
        with contextlib.redirect_stdout(io.StringIO()):   # main() prints a summary
            rc = pw.main(["--person", f"Alpha={ALPHA}", "--person", f"Beta={BETA}",
                          "--input", src, "--out", out])
        self.assertEqual(rc, 0)
        return load_workbook(out)

    def _summary_row(self, ws, label):
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == label:
                return row
        raise AssertionError(f"no summary row named {label!r}")

    def test_measured_zero_renders_as_zero(self):
        # A PR that nobody reviewed. That is a zero, and it is the
        # finding; rendering it as "no data" would bury it.
        wb = self._build([pr(BETA, 1, reviewer_count=0)])
        row = self._summary_row(wb["Summary"], "PRs with at least one reviewer")
        self.assertEqual(row[2], 0)

    def test_absent_source_renders_as_no_data_not_zero(self):
        # Alpha has no pull requests at all in this dataset.
        wb = self._build([pr(BETA, 1)])
        row = self._summary_row(wb["Summary"], "PRs merged")
        self.assertEqual(row[1], "no data")
        self.assertEqual(row[2], 1)

    def test_every_sheet_is_present_even_when_empty(self):
        wb = self._build([pr(BETA, 1)])
        for name in ("Summary", "Jira Issues", "Jira Transitions",
                     "Pull Requests", "Reviews Given", "Reverts",
                     "Coverage & Gaps"):
            self.assertIn(name, wb.sheetnames)

    def test_pr_titles_are_never_written(self):
        # The schema carries no title text; assert the sheet cannot grow one.
        wb = self._build([pr(BETA, 1)])
        headers = [c.value for c in wb["Pull Requests"][1]]
        self.assertNotIn("Title", headers)
        self.assertIn("AI marker", headers)

    def test_only_named_people_appear(self):
        wb = self._build([pr(BETA, 1), pr("outsider", 2)])
        # row[1] is the PR id; the trailing footnote has none, so it filters out.
        names = {row[0] for row in wb["Pull Requests"].iter_rows(
            min_row=2, values_only=True) if row[0] and row[1] is not None}
        self.assertEqual(names, {"Beta"})

    def test_unresolved_issue_is_text_not_a_zero_cycle_time(self):
        wb = self._build([transition(BETA, "PROJ-1", assignee=BETA)])
        row = next(r for r in wb["Jira Issues"].iter_rows(min_row=2,
                                                          values_only=True)
                   if r[1] == "PROJ-1")
        self.assertEqual(row[9], "not resolved")
        self.assertEqual(row[10], "n/a")

    def test_unrecognised_ai_labels_reach_the_report(self):
        """Marker drift is a DQ finding, so it has to be visible, by name.

        PLANNER_BY_COPILOT / COPILOT_TESTING are real names found in live Jira.
        Each is AI work that the AI count is missing; a silent drop makes the
        gap invisible, and a bare boolean cannot tell anyone what to reconcile.
        """
        event = transition(BETA, "PROJ-1", assignee=BETA)
        event["attributes"]["attribution"] = {
            "has_ai_labels": True,
            "unrecognised_ai_labels": ["COPILOT_TESTING", "PLANNER_BY_COPILOT"],
            "has_ai_label_drift": True,
        }
        wb = self._build([event])

        headers = [c.value for c in wb["Jira Issues"][1]]
        self.assertEqual(headers.index("Unrecognised AI labels"), 13)
        row = next(r for r in wb["Jira Issues"].iter_rows(min_row=2, values_only=True)
                   if r[1] == "PROJ-1")
        self.assertIn("PLANNER_BY_COPILOT", row[13])
        self.assertIn("COPILOT_TESTING", row[13])

        summary = self._summary_row(wb["Summary"], "Issues with an UNRECOGNISED AI label")
        self.assertEqual(summary[2], 1)

    def test_drift_free_issue_reports_a_measured_zero_not_no_data(self):
        wb = self._build([transition(BETA, "PROJ-1", assignee=BETA)])
        summary = self._summary_row(wb["Summary"], "Issues with an UNRECOGNISED AI label")
        self.assertEqual(summary[2], 0)

    def test_bad_person_spec_is_rejected(self):
        self.assertEqual(pw.main(["--person", "NoEqualsSign",
                                  "--input", os.devnull, "--out", "x.xlsx"]), 2)


def make_test_run(actor, case_key, cycle="CY-1", status="Passed", category="passed",
                  executed_by=None, assigned_to=None, automated=False, defects=0,
                  when="2026-08-12T10:00:00Z"):
    return base("test.run.completed", actor, when=when, attrs={
        "test_case_key": case_key, "test_cycle_key": cycle,
        "test_run_id": hash(case_key) % 10000, "status": status,
        "status_category": category, "is_automated": automated,
        "executed_by_person_id": executed_by, "assigned_to_person_id": assigned_to,
        "executed_at": None if category == "not_run" else when,
        "effort_seconds": None, "defect_count": defects,
        "folder_name": "Regression", "priority": "High",
    }, event_id=f"evt_tr_{cycle}_{case_key}")


class TestAioAggregation(unittest.TestCase):
    def test_execution_is_credited_to_the_executor_not_the_assignee(self):
        evs = [make_test_run(ALPHA, "TC-1", executed_by=ALPHA, assigned_to=BETA)]
        data = pw.collect(evs, PEOPLE, None)
        self.assertEqual(len(data["Alpha"].test_runs), 1)
        self.assertEqual(len(data["Beta"].test_runs), 0)
        # Beta still sees it, but as an assignment they did not execute.
        self.assertEqual(len(data["Beta"].test_assigned), 1)

    def test_self_assigned_run_is_not_double_counted(self):
        evs = [make_test_run(ALPHA, "TC-1", executed_by=ALPHA, assigned_to=ALPHA)]
        data = pw.collect(evs, PEOPLE, None)
        self.assertEqual(len(data["Alpha"].test_runs), 1)
        self.assertEqual(len(data["Alpha"].test_assigned), 0)

    def test_not_run_is_outside_the_executed_denominator(self):
        evs = [
            make_test_run(ALPHA, f"TC-{i}", executed_by=ALPHA,
                     status="Passed", category="passed") for i in range(3)
        ] + [
            make_test_run(ALPHA, f"TC-nr-{i}", executed_by=ALPHA,
                     status="Not Run", category="not_run") for i in range(50)
        ]
        totals = pw.test_totals(pw.collect(evs, PEOPLE, None)["Alpha"])
        # 50 seeded rows must not drown a 100% pass rate down to 5.7%.
        self.assertEqual(totals["executed"], 3)
        self.assertEqual(totals["passed"], 3)

    def test_pass_rate_keeps_blocked_in_the_denominator(self):
        evs = ([make_test_run(ALPHA, f"TC-p{i}", executed_by=ALPHA) for i in range(9)] +
               [make_test_run(ALPHA, "TC-b", executed_by=ALPHA, status="Blocked",
                         category="blocked")])
        totals = pw.test_totals(pw.collect(evs, PEOPLE, None)["Alpha"])
        self.assertEqual(totals["executed"], 10)
        # A blocked test is one that could not be run; excluding it would report
        # 100% instead of 90% and flatter the result.
        self.assertEqual(totals["passed"], 9)


class TestAioSheets(unittest.TestCase):
    def _wb(self, events):
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, "e.ndjson")
        with open(src, "w") as fh:
            for e in events:
                fh.write(json.dumps(e) + "\n")
        out = os.path.join(tmp, "out.xlsx")
        with contextlib.redirect_stdout(io.StringIO()):
            pw.main(["--person", f"Alpha={ALPHA}", "--person", f"Beta={BETA}",
                     "--input", src, "--out", out])
        return load_workbook(out)

    def test_cycle_rollup_separates_rows_from_executed(self):
        evs = ([make_test_run(ALPHA, f"TC-{i}", executed_by=ALPHA) for i in range(4)] +
               [make_test_run(ALPHA, "TC-x", executed_by=ALPHA, status="Not Run",
                         category="not_run")])
        wb = self._wb(evs)
        row = next(r for r in wb["Test Cycles"].iter_rows(min_row=2, values_only=True)
                   if r[1] == "CY-1")
        self.assertEqual(row[2], 5)   # rows touched
        self.assertEqual(row[3], 4)   # executed
        self.assertEqual(row[8], 100.0)

    def test_test_run_sheet_omits_titles_and_shows_never_run(self):
        wb = self._wb([make_test_run(ALPHA, "TC-1", executed_by=ALPHA,
                                status="Not Run", category="not_run")])
        headers = [c.value for c in wb["Test Runs"][1]]
        self.assertNotIn("Title", headers)
        row = next(r for r in wb["Test Runs"].iter_rows(min_row=2, values_only=True)
                   if r[2] == "TC-1")
        self.assertEqual(row[6], "never run")


class TestTrendCharts(unittest.TestCase):
    """A chart that renders as an empty frame is worse than no chart at all."""

    def _wb(self, events):
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, "e.ndjson")
        with open(src, "w") as fh:
            for e in events:
                fh.write(json.dumps(e) + "\n")
        out = os.path.join(tmp, "out.xlsx")
        with contextlib.redirect_stdout(io.StringIO()):
            pw.main(["--person", f"Alpha={ALPHA}", "--person", f"Beta={BETA}",
                     "--input", src, "--out", out])
        return load_workbook(out), out

    def _events(self):
        return [
            pr(ALPHA, 1, merged_at="2026-07-06T10:00:00Z"),
            pr(BETA, 2, merged_at="2026-07-13T10:00:00Z"),
            make_test_run(ALPHA, "TC-1", executed_by=ALPHA, when="2026-07-06T10:00:00Z"),
            make_test_run(BETA, "TC-2", executed_by=BETA, when="2026-07-13T10:00:00Z"),
        ]

    def test_charts_are_created(self):
        wb, _ = self._wb(self._events())
        self.assertIn("Trend Charts", wb.sheetnames)
        self.assertEqual(len(wb["Trend Charts"]._charts), len(pw.TREND_METRICS))

    def test_chart_source_columns_are_not_hidden(self):
        # Excel writes plotVisOnly=1 by default, so a chart whose source columns
        # are hidden renders as an empty frame. This is the bug that shipped once.
        wb, _ = self._wb(self._events())
        for sheet in wb.sheetnames:
            for key, dim in wb[sheet].column_dimensions.items():
                self.assertFalse(dim.hidden,
                                 f"{sheet}!{key} is hidden; charts would go blank")

    def test_plot_vis_only_is_disabled(self):
        # Asserted on the written XML, not the in-memory object: openpyxl's
        # chart *reader* does not surface plotVisOnly, and what Excel obeys is
        # the file on disk.
        _wb, path = self._wb(self._events())
        import zipfile
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("xl/charts/chart1.xml").decode("utf-8")
        self.assertIn('plotVisOnly val="0"', xml)

    def test_week_labels_are_a_string_reference_not_a_number_reference(self):
        # ISO week labels are text. As a numRef Excel may render the axis as
        # 1, 2, 3 instead of 2026-W28.
        _wb, path = self._wb(self._events())
        import zipfile
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("xl/charts/chart1.xml").decode("utf-8")
        self.assertIn("<cat><strRef>", xml)
        self.assertNotIn("<cat><numRef>", xml)

    def test_series_reference_a_sheet_that_holds_numbers(self):
        wb, _ = self._wb(self._events())
        chart = wb["Trend Charts"]._charts[0]
        ref = str(chart.series[0].val.numRef.f)
        sheet_name = ref.split("!")[0].strip("'")
        self.assertIn(sheet_name, wb.sheetnames)
        # And the referenced block must actually contain a value, not be empty.
        values = [c.value for row in wb[sheet_name].iter_rows() for c in row]
        self.assertTrue(any(isinstance(v, (int, float)) for v in values))

    def test_partial_week_is_excluded_from_charts_but_kept_in_the_table(self):
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        wb, _ = self._wb(self._events() + [pr(ALPHA, 3, merged_at=now)])
        current = pw.iso_week(datetime.now(timezone.utc))
        table_weeks = {str(r[1]) for r in wb["Weekly Trend"].iter_rows(
            min_row=2, values_only=True) if r[1]}
        self.assertIn(f"{current} (partial)", table_weeks)
        chart_weeks = {c.value for row in wb["Trend Data"].iter_rows()
                       for c in row if isinstance(c.value, str)
                       and c.value.startswith("20") and "-W" in c.value}
        self.assertNotIn(current, chart_weeks)


if __name__ == "__main__":
    unittest.main(verbosity=2)
