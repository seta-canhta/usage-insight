#!/usr/bin/env python3
"""Tests for the AI-usage half of the per-person workbook.

Most of these are about denominators. Every wrong number this file has produced
came from one -- coverage read over the case estate instead of the cycle, a
partial week compared against a full one, an unpriced model treated as free.
"""

import json
import os
import sys
import tempfile
import unittest

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPORT = os.path.dirname(_HERE)
for _p in (_REPORT, os.path.dirname(_REPORT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import ai_usage_sheets as sheets  # noqa: E402

NGOC = "5bee6a1ec03ef4570f0a78e3"
LINH = "712020:28cc987e-5263-4564-83c6-7f76fa32574e"
PEOPLE = {NGOC: "Ngoc Nguyen", LINH: "Linh Hoang"}
WEEKS = ["2026-W32", "2026-W33", "2026-W34"]
PRICES = {"claude-sonnet-4.6": (3.0, 15.0)}


def event(kind, when, who=NGOC, **attrs):
    return {"event_type": kind, "event_time": when,
            "actor": {"person_id": who}, "trace_id": attrs.pop("trace", "s1"),
            "attributes": attrs}


class ArgParsingTests(unittest.TestCase):
    def test_a_week_span_expands_inclusively(self):
        self.assertEqual(sheets.week_span("2026-W31..2026-W35"),
                         ["2026-W31", "2026-W32", "2026-W33",
                          "2026-W34", "2026-W35"])

    def test_a_single_week_is_a_span_of_one(self):
        self.assertEqual(sheets.week_span("2026-W34"), ["2026-W34"])

    def test_a_price_is_input_slash_output_per_million(self):
        self.assertEqual(sheets.parse_price("claude-opus-4.6=5.0/25.0"),
                         ("claude-opus-4.6", (5.0, 25.0)))

    def test_a_malformed_price_is_refused_rather_than_defaulted(self):
        """A price that silently became 0.0 would report AI as free."""
        with self.assertRaises(Exception):
            sheets.parse_price("claude-opus-4.6=free")

    def test_a_person_needs_both_halves(self):
        with self.assertRaises(Exception):
            sheets.parse_person("Ngoc Nguyen")


class CoverageByCycleTests(unittest.TestCase):
    """Metric 2's denominator is the cycle, not the case estate."""

    def write(self, rows):
        fh = tempfile.NamedTemporaryFile("w", suffix=".ndjson", delete=False,
                                         encoding="utf-8")
        for r in rows:
            fh.write(json.dumps(r) + "\n")
        fh.close()
        self.addCleanup(os.unlink, fh.name)
        return fh.name

    def run_event(self, cycle, case, automated):
        return {"event_type": "test.run.completed",
                "event_time": "2026-08-12T09:00:00Z",
                "actor": {"person_id": NGOC},
                "attributes": {"test_cycle_key": cycle, "test_case_key": case,
                               "is_automated": automated,
                               "executed_at": "2026-08-12T09:00:00Z",
                               "status_category": "passed",
                               "folder_name": "Regression"}}

    def test_coverage_is_per_cycle_and_counts_distinct_cases(self):
        path = self.write([
            self.run_event("CY-1", "TC-1", True),
            self.run_event("CY-1", "TC-1", True),      # same case again
            self.run_event("CY-1", "TC-2", False),
            self.run_event("CY-2", "TC-3", False),
        ])
        cov = sheets.collect([path], PEOPLE, WEEKS, PRICES)["coverage_by_cycle"]
        self.assertEqual(cov["CY-1"]["cases"], 2)
        self.assertEqual(cov["CY-1"]["automated"], 1)
        self.assertEqual(cov["CY-1"]["pct"], 50.0)
        self.assertEqual(cov["CY-2"]["pct"], 0.0)

    def test_an_untriaged_estate_case_never_enters_a_cycle_denominator(self):
        """The whole point: a case in no cycle is backlog, not coverage.

        Measured 2026-08-27 on IML -- 3,695 of 5,183 P3 cases carry no
        automation status and sit in no cycle. Counting them as 'not automated'
        put P3 at 22.8% and read as a delivery crisis.
        """
        path = self.write([
            self.run_event("CY-1", "TC-1", True),
            {"event_type": "test.case.snapshot",
             "event_time": "2026-08-12T09:00:00Z", "actor": {},
             "attributes": {"test_case_key": "TC-99", "priority": "Low",
                            "automation_status": None}},
        ])
        data = sheets.collect([path], PEOPLE, WEEKS, PRICES)
        self.assertEqual(data["coverage_by_cycle"]["CY-1"]["cases"], 1)
        self.assertEqual(data["coverage_by_cycle"]["CY-1"]["pct"], 100.0)
        # ...and it is still counted, separately, as backlog.
        self.assertEqual(data["estate"]["Low"]["unset"], 1)

    def test_a_fully_manual_cycle_is_zero_percent_not_missing(self):
        path = self.write([self.run_event("CY-M", "TC-1", False)])
        cov = sheets.collect([path], PEOPLE, WEEKS, PRICES)["coverage_by_cycle"]
        self.assertEqual(cov["CY-M"]["pct"], 0.0)


class CostTests(unittest.TestCase):
    def collect(self, events, prices=PRICES):
        fh = tempfile.NamedTemporaryFile("w", suffix=".ndjson", delete=False,
                                         encoding="utf-8")
        for e in events:
            fh.write(json.dumps(e) + "\n")
        fh.close()
        self.addCleanup(os.unlink, fh.name)
        return sheets.collect([fh.name], PEOPLE, WEEKS, prices)["cost"]

    def test_an_unpriced_model_is_not_free(self):
        """No --price for a model means unpriced, never zero."""
        cost = self.collect([
            event("model.call", "2026-08-12T09:00:00Z",
                  input_tokens=1000, output_tokens=100, model_id="mystery-1"),
        ])
        row = cost["Ngoc Nguyen"]["2026-W33"]
        self.assertIsNone(row["modelled"])
        self.assertEqual(row["calls"], 1)
        self.assertEqual(row["unpriced"], 1)

    def test_calls_without_tokens_are_projected_from_the_priced_median(self):
        priced = event("model.call", "2026-08-12T09:00:00Z",
                       input_tokens=1_000_000, output_tokens=0,
                       model_id="claude-sonnet-4.6")
        bare = event("model.call", "2026-08-12T10:00:00Z",
                     model_id="claude-sonnet-4.6")
        row = self.collect([priced, bare])["Ngoc Nguyen"]["2026-W33"]
        self.assertEqual(row["measured"], 3.0)   # 1M input at $3/1M
        self.assertEqual(row["modelled"], 6.0)   # the bare call projected
        self.assertEqual(row["n"], 1)

    def test_a_week_with_no_calls_is_none_rather_than_zero(self):
        cost = self.collect([event("human.turn", "2026-08-12T09:00:00Z")])
        self.assertIsNone(cost["Ngoc Nguyen"]["2026-W32"])


class ChangeColumnTests(unittest.TestCase):
    """A partial week must never set the trend."""

    def test_change_spans_full_weeks_only(self):
        import openpyxl
        wb = openpyxl.Workbook()
        weeks = ["2026-W32", "2026-W33", "2026-W34", "2026-W35"]
        data = sheets.collect([], PEOPLE, weeks, PRICES)
        sheets.render(wb, data, weeks, ["2026-W32", "2026-W34"],
                      {"2026-W35": "Mon-Thu only"})
        ws = wb["Ten Metrics"]
        self.assertEqual(ws.cell(row=1, column=4 + len(weeks) + 1).value,
                         "W32->W34")

    def test_an_empty_pull_renders_rather_than_dividing_by_zero(self):
        """No test-cycle events must not crash, and must not read as 0%.

        Absent is not zero -- "0% automated" is the most damaging possible way
        to render a source that simply was not pulled.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        weeks = ["2026-W32", "2026-W33"]
        data = sheets.collect([], PEOPLE, weeks, PRICES)
        sheets.render(wb, data, weeks, weeks, {})
        text = " ".join(str(c.value) for row in wb["Start Here"].iter_rows()
                        for c in row if c.value)
        self.assertIn("not measured", text)
        self.assertNotIn("0.0%", text)

    def test_full_weeks_outside_the_span_are_refused(self):
        with self.assertRaises(SystemExit):
            sheets.main(["nonexistent.xlsx", "--person", "A=1",
                         "--input", ".", "--weeks", "2026-W32..2026-W33",
                         "--full-weeks", "2026-W40"])


if __name__ == "__main__":
    unittest.main()
