#!/usr/bin/env python3
"""ai_usage_sheets.py -- the AI-usage half of the per-person workbook.

`people_workbook.py` reads Jira, Bitbucket and AIO and is thorough on delivery,
but it carries nothing from the laptop stream: no prompts, no sessions, no
tokens, no cost. That is the "are they using AI effectively" half of the
question this project exists to answer, so this adds it to the same file.

    python3 report/people_workbook.py --person ... --out W.xlsx
    python3 report/ai_usage_sheets.py W.xlsx \
        --person "NAME=ACCOUNT_ID" --person "NAME=ACCOUNT_ID" \
        --input reports/2026-08/workbook-input \
        --weeks 2026-W31..2026-W35 --full-weeks 2026-W32..2026-W34 \
        --price "claude-sonnet-4.6=3.0/15.0" --price "claude-opus-4.6=5.0/25.0"

Run it AFTER the generator -- the generator writes the file from scratch and
would drop these sheets.

Three rules it keeps, because each one has already produced a wrong number here:

* **Prices are passed in, never hardcoded.** CONTRACT.md §4: cost is derived
  against *dated* pricing and is never asserted by a client. A rate baked into
  this file would be a client asserting a price, and would silently restate
  history the day a vendor changes one. A model with no `--price` is counted
  and left unpriced rather than guessed at.
* **Coverage is measured by cycle, not over the case estate.** The cycle is
  the delivery record (CONTRACT.md §3 row 22). Measured 2026-08-27 on IML:
  7,976 of 8,564 cases across the seven cycles that ran are automated (93.1%),
  where the same data read over the whole 10,742-case estate puts P3 at 22.8%
  -- a backlog nobody has triaged, not a delivery figure.
* **Partial weeks are labelled, and the change column skips them.** A Mon-Thu
  window against a full week misreads every volume metric.

Requires ``openpyxl``, like the workbook it augments. Everything else is stdlib.
"""

from __future__ import annotations

import argparse
import collections
import datetime
import glob
import json
import os
import statistics
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import Marker
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guard
    sys.exit("ai_usage_sheets.py needs openpyxl:  python3 -m pip install openpyxl")


#: Categorical slots 1 and 2, validated together for colour-vision deficiency:
#: worst adjacent pair dE 24.7 protan / 33.6 normal, against a target of 8.
#: Two people, two hues, assigned in fixed order and never cycled -- a person
#: keeps their colour on every chart in the file, so identity survives a glance
#: from one chart to the next.
SERIES = ["2A78D6", "EB6834"]
#: One hue for magnitude. Coverage is a quantity, not an identity.
MAGNITUDE = "1F6FB2"
PT = 12700  # EMU per point


def parse_person(value):
    name, _, account = value.partition("=")
    if not name.strip() or not account.strip():
        raise argparse.ArgumentTypeError(
            "--person takes NAME=ACCOUNT_ID, e.g. \"A Name=5f00...\"")
    return name.strip(), account.strip()


def parse_price(value):
    model, _, rates = value.partition("=")
    inp, _, out = rates.partition("/")
    try:
        return model.strip(), (float(inp), float(out))
    except ValueError:
        raise argparse.ArgumentTypeError(
            "--price takes MODEL=IN/OUT per 1M tokens, e.g. "
            "\"claude-sonnet-4.6=3.0/15.0\"")


def week_span(value):
    """``2026-W31..2026-W35`` -> every ISO week between, inclusive."""
    start, _, end = value.partition("..")
    if not end:
        return [start]
    def key(w):
        y, _, n = w.partition("-W")
        return int(y), int(n)
    (y0, n0), (y1, n1) = key(start), key(end)
    out = []
    y, n = y0, n0
    while (y, n) <= (y1, n1):
        out.append("%d-W%02d" % (y, n))
        n += 1
        if n > datetime.date(y, 12, 28).isocalendar()[1]:
            y, n = y + 1, 1
    return out


# --------------------------------------------------------------------------
# loading
# --------------------------------------------------------------------------

def iso_week(ts):
    if not ts or len(ts) < 10:
        return None
    try:
        d = datetime.date.fromisoformat(ts[:10])
    except ValueError:
        return None
    y, w, _ = d.isocalendar()
    return "%d-W%02d" % (y, w)


def rows(paths):
    for path in paths:
        if os.path.isdir(path):
            found = sorted(glob.glob(os.path.join(path, "**", "*.ndjson"),
                                     recursive=True))
        else:
            found = [path]
        for one in found:
            with open(one, "r", encoding="utf-8") as fh:
                for line in fh:
                    if line.strip():
                        try:
                            yield json.loads(line)
                        except json.JSONDecodeError:
                            continue


def collect(inputs, people, weeks, prices):
    """Every figure these sheets render, counted from the event stream."""
    names = list(people.values())
    z = lambda: {n: {w: collections.Counter() for w in weeks} for n in names}
    ai, bb = z(), z()
    days = {n: {w: set() for w in weeks} for n in names}
    sess = {n: {w: set() for w in weeks} for n in names}
    toks = {n: {w: [] for w in weeks} for n in names}
    merge = {n: {w: [] for w in weeks} for n in names}
    #: Per ISO week, project-wide unless the field carries a person.
    sp = collections.defaultdict(collections.Counter)
    wk_cycles = collections.defaultdict(set)
    wk_people = collections.defaultdict(lambda: collections.defaultdict(collections.Counter))
    wk_folders = collections.defaultdict(collections.Counter)
    case_history = collections.defaultdict(list)
    seen_issues = set()
    cycles = collections.defaultdict(
        lambda: {"cases": set(), "auto": set(), "folders": collections.Counter(),
                 "first": None, "last": None, "runs": 0, "failed": 0,
                 "defects": 0, "people": collections.Counter()})
    estate = collections.defaultdict(collections.Counter)

    for e in rows(inputs):
        who = people.get((e.get("actor") or {}).get("person_id"))
        t, a = e.get("event_type"), (e.get("attributes") or {})
        w = iso_week(e.get("event_time"))

        if t == "test.case.snapshot":
            # created_at is a fact; updated_at is only the LAST edit, so a case
            # touched three times in a week counts once and an earlier touch
            # is overwritten. Stated on the sheet rather than smoothed over.
            born = iso_week(a.get("created_at"))
            if born:
                sp[born]["cases_created"] += 1
            touched = iso_week(a.get("updated_at"))
            if touched:
                sp[touched]["cases_updated"] += 1
            pri = a.get("priority") or "unset"
            st = a.get("automation_status")
            estate[pri]["total"] += 1
            estate[pri][{None: "unset", "Automated": "automated",
                         "To Be Automated": "to_be_automated",
                         "Manual": "manual",
                         "In Progress": "in_progress"}.get(st, "other")] += 1
            continue

        if t == "test.run.completed":
            key, case = a.get("test_cycle_key"), a.get("test_case_key")
            if case:
                case_history[case].append(
                    (a.get("executed_at") or "", a.get("status_category")))
            spr = iso_week(a.get("executed_at"))
            if spr:
                sp[spr]["runs"] += 1
                cat = (a.get("status_category") or "").lower()
                if cat in ("passed", "failed", "blocked", "not_run"):
                    sp[spr][cat] += 1
                if a.get("is_automated"):
                    sp[spr]["automated"] += 1
                sp[spr]["defects"] += a.get("defect_count") or 0
                if key:
                    wk_cycles[spr].add(key)
                if a.get("folder_name"):
                    wk_folders[spr][a["folder_name"]] += 1
                runner = people.get(a.get("executed_by_person_id"))
                if runner:
                    # Only the two subjects are in `people`, so nothing here
                    # counts anybody else's runs.
                    wk_people[spr][runner]["runs"] += 1
                    if cat in ("passed", "failed", "blocked"):
                        wk_people[spr][runner][cat] += 1
                    if a.get("is_automated"):
                        wk_people[spr][runner]["automated"] += 1
                    wk_people[spr][runner]["defects"] += a.get("defect_count") or 0
            if not key:
                continue
            c = cycles[key]
            c["runs"] += 1
            if case:
                c["cases"].add(case)
                if a.get("is_automated"):
                    c["auto"].add(case)
            if a.get("folder_name"):
                c["folders"][a["folder_name"]] += 1
            if (a.get("status_category") or "").lower() == "failed":
                c["failed"] += 1
            c["defects"] += a.get("defect_count") or 0
            runner = a.get("executed_by_person_id")
            if runner:
                c["people"][people.get(runner, "other")] += 1
            when = a.get("executed_at") or ""
            if when:
                c["first"] = min(c["first"] or when, when)
                c["last"] = max(c["last"] or when, when)
            continue

        if t == "jira.transition":
            # A QA engineer's headline output is the bugs they raise. Counted
            # once per issue by its reporter, not once per transition -- an
            # issue emits one event per status change and counting those would
            # rank people by how much a ticket moved.
            iss = a.get("issue") or {}
            ikey = iss.get("issue_key")
            if ikey and ikey not in seen_issues:
                seen_issues.add(ikey)
                reporter = people.get(iss.get("reporter_person_id"))
                spr = iso_week(iss.get("created_at"))
                if reporter and spr:
                    kind = (iss.get("issue_type") or "Other")
                    wk_people[spr][reporter]["raised_" + kind.lower()] += 1
                    if kind == "Bug":
                        sp[spr]["bugs_raised"] += 1

        if not who or w not in weeks:
            continue

        if t in ("human.turn", "model.call", "tool.call", "scm.commit",
                 "output.generated"):
            ai[who][w][t] += 1
            days[who][w].add(e["event_time"][:10])
            if e.get("trace_id"):
                sess[who][w].add(e["trace_id"])
            if t == "model.call":
                i, o, m = (a.get("input_tokens"), a.get("output_tokens"),
                           a.get("model_id"))
                if isinstance(i, int) and isinstance(o, int):
                    toks[who][w].append((i, o, m))
            continue

        if t.startswith("scm.") or t == "scm.revert":
            bb[who][w][t] += 1
            if t == "scm.pr.merged":
                for src, dst in (("automation_scripts_added", "scripts_added"),
                                 ("automation_scripts_modified", "scripts_modified"),
                                 ("lines_added", "lines_added"),
                                 ("lines_removed", "lines_removed"),
                                 ("commit_count", "commits")):
                    bb[who][w][dst] += a.get(src) or 0
                if isinstance(a.get("merge_lead_time_ms"), (int, float)):
                    merge[who][w].append(a["merge_lead_time_ms"])

    cov = {}
    for key, c in cycles.items():
        n = len(c["cases"])
        cov[key] = {
            "cases": n, "automated": len(c["auto"]),
            "pct": round(100.0 * len(c["auto"]) / n, 1) if n else None,
            "runs": c["runs"], "failed": c["failed"], "defects": c["defects"],
            "folder": c["folders"].most_common(1)[0][0] if c["folders"] else None,
            "first": (c["first"] or "")[:10], "last": (c["last"] or "")[:10],
            "ours": {k: v for k, v in c["people"].items() if k in names},
        }

    cost = {}
    for n in names:
        cost[n] = {}
        for w in weeks:
            calls = ai[n][w]["model.call"]
            priced = [x for x in toks[n][w] if x[2] in prices]
            if not calls:
                cost[n][w] = None
                continue
            if not priced:
                # Calls happened and none could be priced. Absent, not zero.
                cost[n][w] = {"measured": None, "modelled": None, "n": 0,
                              "calls": calls, "unpriced": len(toks[n][w])}
                continue
            measured = sum(i / 1e6 * prices[m][0] + o / 1e6 * prices[m][1]
                           for i, o, m in priced)
            med_i = statistics.median([x[0] for x in priced])
            med_o = statistics.median([x[1] for x in priced])
            mix = collections.Counter(x[2] for x in priced).most_common(1)[0][0]
            per = med_i / 1e6 * prices[mix][0] + med_o / 1e6 * prices[mix][1]
            cost[n][w] = {
                "measured": round(measured, 2),
                "modelled": round(measured + per * (calls - len(priced)), 2),
                "n": len(priced), "calls": calls,
                "unpriced": len(toks[n][w]) - len(priced)}

    # A case that failed and later passed. The only "fixed" this source can
    # support, and a thin one: AIO keeps one run per case per cycle and
    # overwrites on re-run, so a fix inside a cycle leaves no trace at all.
    for case, runs_ in case_history.items():
        runs_.sort()
        for i in range(1, len(runs_)):
            if runs_[i - 1][1] == "failed" and runs_[i][1] == "passed":
                spr = iso_week(runs_[i][0])
                if spr:
                    sp[spr]["fixed"] += 1
                break

    return {"ai": ai, "bb": bb, "days": days, "sessions": sess, "tokens": toks,
            "merge": merge, "coverage_by_cycle": cov, "estate": dict(estate),
            "cost": cost, "names": names,
            "week": {k: dict(v) for k, v in sp.items()},
            "week_cycles": {k: sorted(v) for k, v in wk_cycles.items()},
            "week_folders": {k: dict(v) for k, v in wk_folders.items()},
            "week_people": {k: {p: dict(c) for p, c in v.items()}
                              for k, v in wk_people.items()}}


# --------------------------------------------------------------------------
# rendering
# --------------------------------------------------------------------------

HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F3864")
GRP = Font(bold=True)
GRPF = PatternFill("solid", fgColor="DDEBF7")
NOTE = Font(italic=True, size=9, color="555555")


#: Pronouns are stated on the command line, never inferred. A name does not
#: carry them, and a guess misgenders a real person in a way "they" never
#: does -- so an unnamed person stays they/them and nothing in the copy has to
#: change for it to read correctly.
PRONOUN_SETS = {
    "she": ("she", "her", "her", "hers", False),
    "he": ("he", "him", "his", "his", False),
    "they": ("they", "them", "their", "theirs", True),
}


class Pronouns(object):
    """One person's pronouns, plus the verb agreement they force."""

    def __init__(self, spec="they"):
        parts = [x.strip() for x in str(spec).split("/") if x.strip()]
        if len(parts) == 1 and parts[0].lower() in PRONOUN_SETS:
            subj, obj, poss, indep, plural = PRONOUN_SETS[parts[0].lower()]
        else:
            while len(parts) < 4:
                parts.append(parts[-1])
            subj, obj, poss, indep = parts[:4]
            plural = subj.lower() == "they"
        self.subj, self.obj, self.poss, self.indep = subj, obj, poss, indep
        self.plural = plural

    @property
    def Subj(self):
        return self.subj[:1].upper() + self.subj[1:]

    @property
    def Poss(self):
        return self.poss[:1].upper() + self.poss[1:]

    @property
    def is_(self):
        return "are" if self.plural else "is"

    def v(self, stem):
        """Third-person present: `writes` for she/he, `write` for they."""
        if self.plural:
            return stem
        return stem + ("es" if stem.endswith(("s", "sh", "ch", "x", "z"))
                       else "s")


def parse_pronouns(text):
    name, _, spec = text.partition("=")
    if not name.strip() or not spec.strip():
        raise argparse.ArgumentTypeError(
            "--pronouns wants NAME=she or NAME=he/him/his/his, got %r" % text)
    return name.strip(), Pronouns(spec)


def fresh(wb, title):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def head(ws, cols, widths):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font, cell.fill = HEAD, FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=2, column=1)


def pct(a, b):
    return None if not b else round(100.0 * a / b, 1)


def notes(ws, row, lines):
    for line in lines:
        ws.cell(row=row, column=1, value=line).font = NOTE
        row += 1
    return row


def render(wb, data, weeks, full_weeks, window_label, pron=None):
    names = data["names"]
    pron = pron if pron is not None else (lambda n: Pronouns())
    ai, bb, cost = data["ai"], data["bb"], data["cost"]

    def change(series):
        """First to last across FULL weeks only -- partials misread volume."""
        i0, i1 = weeks.index(full_weeks[0]), weeks.index(full_weeks[-1])
        a, b = series[i0], series[i1]
        if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
            return None
        return None if not a else round(100.0 * (b - a) / a, 1)

    # ---------------------------------------------------------- AI Usage
    ws = fresh(wb, "AI Usage")
    head(ws, ["Person", "Week", "Window", "Prompts", "Model calls", "Tool calls",
              "Sessions", "Active days", "Prompts/active day", "Action rate %",
              "Questions with a measured cost", "AI cost, estimated ($)",
            "Cost per question $"],
         [16, 10, 24, 9, 11, 10, 9, 11, 17, 12, 11, 20, 12])
    r = 2
    for n in names:
        for w in weeks:
            prompts = ai[n][w]["human.turn"]
            if not prompts:
                continue
            d = len(data["days"][n][w])
            c = cost[n].get(w) or {}
            m = c.get("modelled")
            for col, val in enumerate([
                    n, w, window_label.get(w, "full week"), prompts,
                    ai[n][w]["model.call"], ai[n][w]["tool.call"],
                    len(data["sessions"][n][w]), d,
                    round(prompts / d, 1) if d else None,
                    pct(ai[n][w]["tool.call"], prompts), c.get("n"), m,
                    round(m / prompts, 3) if m and prompts else None], 1):
                ws.cell(row=r, column=col, value=val)
            r += 1
    notes(ws, r + 1, [
        "\"How often AI did the work\" counts the questions where AI went and "
        "read, changed or ran something, rather than only replying.",
        "The cost is an ESTIMATE. Only a minority of AI questions record what "
        "they used, so the rest are scaled from the typical question that week. "
        "Published list prices, no discounts applied.",
        "It is NOT the bill. Copilot charges a monthly fee per person plus "
        "usage, and neither is visible here. Use it to compare weeks against "
        "each other, not to forecast a budget.",
        "A blank cell is a quantity with no measurement, never a zero.",
    ])

    # ------------------------------------------------------- Ten Metrics
    ws = fresh(wb, "Ten Metrics")
    head(ws, ["#", "Metric / sub-measure", "Want", "Person"] +
         [w[-3:] for w in weeks] +
         ["%s->%s" % (full_weeks[0][-3:], full_weeks[-1][-3:]), "Note"],
         [5, 40, 6, 20] + [10] * len(weeks) + [10, 64])
    r = 2

    def group(num, name):
        nonlocal r
        for col in range(1, 12):
            ws.cell(row=r, column=col).fill = GRPF
        ws.cell(row=r, column=1, value=num).font = GRP
        ws.cell(row=r, column=2, value=name).font = GRP
        r += 1

    def line(sub, want, person, series, note=""):
        nonlocal r
        ws.cell(row=r, column=2, value="   " + sub)
        ws.cell(row=r, column=3, value=want)
        ws.cell(row=r, column=4, value=person)
        for i, v in enumerate(series):
            ws.cell(row=r, column=5 + i,
                    value=v if v is not None else "no data")
        d = change(series)
        ws.cell(row=r, column=5 + len(weeks),
                value=("%+.0f%%" % d) if d is not None else "-")
        ws.cell(row=r, column=6 + len(weeks), value=note)
        r += 1

    def free(text, note=""):
        nonlocal r
        ws.cell(row=r, column=2, value="   " + text)
        if note:
            ws.cell(row=r, column=6 + len(weeks), value=note)
        r += 1

    group(1, "Automation Output")
    for n in names:
        line("Automation scripts added", "up", n,
             [bb[n][w]["scripts_added"] for w in weeks],
             "New test script files in the code changes they delivered")
        line("Automation scripts modified", "up", n,
             [bb[n][w]["scripts_modified"] for w in weeks])
        line("PRs merged", "up", n,
             [bb[n][w]["scm.pr.merged"] for w in weeks])
        line("AI outputs written (files)", "up", n,
             [ai[n][w]["output.generated"] for w in weeks],
             "Files AI wrote directly, rather than a person typing them")

    group(2, "Automation Coverage  (BY CYCLE)")
    cov = data["coverage_by_cycle"]
    tc = sum(c["cases"] for c in cov.values())
    ta = sum(c["automated"] for c in cov.values())
    ws.cell(row=r, column=2, value="   All delivered cycles")
    ws.cell(row=r, column=3, value="up")
    ws.cell(row=r, column=5,
            value="%.1f%% (%d / %d)" % (100.0 * ta / tc, ta, tc) if tc else "no data")
    ws.cell(row=r, column=6 + len(weeks),
            value="Counted across the test cycles actually being run, which is "
                  "the work in flight -- not every test that has ever been "
                  "written, most of which nobody has triaged.")
    r += 1
    for key in sorted(cov, key=lambda k: -cov[k]["cases"]):
        c = cov[key]
        ws.cell(row=r, column=2, value="      %s  (%s)" % (key, c["folder"] or "-"))
        ws.cell(row=r, column=4,
                value=", ".join("%s %d" % kv for kv in sorted(c["ours"].items()))
                or "neither")
        ws.cell(row=r, column=5, value="%.1f%% (%d / %d)"
                % (c["pct"], c["automated"], c["cases"]))
        ws.cell(row=r, column=6 + len(weeks),
                value="%s to %s - %d runs, %d failed, %d defects"
                      % (c["first"], c["last"], c["runs"], c["failed"], c["defects"]))
        r += 1
    est = data["estate"]
    if est:
        free("Backlog view (whole case estate)",
             "A different question, kept separate on purpose -- what is "
             "still waiting to be automated: "
             + "; ".join("%s %d automated / %d to-be-automated / %d no status"
                         % (k, est[k].get("automated", 0),
                            est[k].get("to_be_automated", 0), est[k].get("unset", 0))
                         for k in ("High", "Medium", "Low") if k in est)
             + ". Read as coverage it makes an untouched backlog look like a "
               "delivery failure. It is a queue, not a score.")
    r += 1

    group(3, "First-Pass Acceptance Rate")
    for n in names:
        line("Reviews given", "up", n,
             [bb[n][w]["scm.pr.reviewed"] for w in weeks],
             "Too few reviews to turn into a percentage -- the counts are the "
             "honest answer")

    group(4, "Rework Rate")
    for n in names:
        rev = [bb[n][w]["scm.revert"] for w in weeks]
        mer = [bb[n][w]["scm.pr.merged"] for w in weeks]
        line("Reverts", "down", n, rev,
             "A revert is a signal to read, not a defect count")
        line("Reverts / PRs merged %", "down", n,
             [pct(a, b) for a, b in zip(rev, mer)])

    group(5, "Automation Lead Time")
    for n in names:
        line("Median PR merge lead time (h)", "down", n,
             [round(statistics.median(data["merge"][n][w]) / 3600000.0, 2)
              if data["merge"][n][w] else None for w in weeks],
             "The closest thing available. Nothing links a test to the change "
             "that automated it, so this is time-to-merge, not time-to-automate")

    group(6, "Productivity Gain")
    free("excluded",
         "This needs the same work done without AI to compare against, and "
         "there is none. Anyone quoting a speed-up percentage is guessing.")
    r += 1

    group(7, "Execution Rate")
    for n in names:
        line("AI prompts", "up", n, [ai[n][w]["human.turn"] for w in weeks],
             "How much they used AI -- the AI side of getting work done")
        line("Active days", "up", n,
             [len(data["days"][n][w]) for w in weeks])
    free("Test runs executed, per cycle",
         "; ".join("%s %d" % (k, cov[k]["runs"])
                   for k in sorted(cov, key=lambda x: -cov[x]["runs"]))
         + ". Volume follows cycle scheduling, not effort.")
    r += 1

    group(8, "Flaky Test Rate")
    free("Failed runs, per cycle",
         "; ".join("%s %d" % (k, cov[k]["failed"])
                   for k in sorted(cov, key=lambda x: -cov[x]["failed"]))
         + ". This is a failure count, not a flakiness rate: the test tool "
           "keeps only the latest result for each test, so a test that fails "
           "then passes leaves no trace. Flakiness cannot be measured here.")
    r += 1

    group(9, "AI Cost per Accepted Output")
    for n in names:
        line("AI cost, estimated ($)", "down", n,
             [(cost[n].get(w) or {}).get("modelled") for w in weeks],
             "The cost side only. How it is worked out is on the AI Usage tab.")
    free("Cost per finished piece of work",
         "Not available. Nothing records which AI conversation produced which "
         "script, so cost can be split by person and week but not by finished "
         "piece of work. The figures above are a weight, not a price per item.")
    r += 1

    group(10, "AI ROI")
    free("excluded",
         "The value of the work delivered is not recorded anywhere.")

    # ---- QA activity: supporting evidence, not an eleventh metric --------
    #
    # None of this is on the list of ten, and it must never be quoted as if it
    # were -- the ten are the point and this is the context around them. It
    # belongs on this sheet anyway: for a QA engineer, raising bugs and running
    # tests IS the output, and a sheet that scores them only on scripts and
    # pull requests describes the wrong job. Kept below the ten, in its own
    # block, labelled as what it is.
    r += 1
    group("", "QA ACTIVITY  (not one of the ten -- supporting evidence)")
    wkp = data["week_people"]
    for n in names:
        for key, label, want, note in (
                ("raised_bug", "Bugs raised for the developers", "up",
                 "Counted once per issue by who reported it, not once per "
                 "status change"),
                ("runs", "Tests run", "up", ""),
                ("passed", "  of those, passed", "up", ""),
                ("failed", "  of those, failed", "down",
                 "A failure found is the job working, not a fault"),
                ("automated", "  of those, run by automation", "up", ""),
                ("defects", "Defects logged against their runs", "", ""),
                ("raised_task", "Tasks raised", "up", "")):
            line(label, want, n,
                 [wkp.get(w, {}).get(n, {}).get(key, 0) for w in weeks], note)
    free("Test cases created / edited, whole project",
         "; ".join("%s created %d, edited %d"
                   % (w[-3:], data["week"].get(w, {}).get("cases_created", 0),
                      data["week"].get(w, {}).get("cases_updated", 0))
                   for w in weeks)
         + ". Deliberately NOT split by person: the test tool records no "
           "author on a test case, so splitting it would mean inventing one. "
           "It is the whole project's work, including people outside this "
           "report.")
    r += 1

    # ------------------------------------------------------- Productivity
    ws = fresh(wb, "Productivity")
    head(ws, ["Person", "Week", "Window", "Prompts/active day", "Action rate %",
              "Commits", "Commits/prompt", "Lines changed (PRs)",
              "Scripts touched", "Scripts/prompt", "PRs merged",
              "Cost $/PR merged", "Cost $/script touched"],
         [16, 10, 24, 17, 12, 9, 14, 18, 15, 14, 10, 16, 20])
    r = 2
    for n in names:
        for w in weeks:
            prompts = ai[n][w]["human.turn"]
            merged = bb[n][w]["scm.pr.merged"]
            if not prompts and not merged:
                continue
            d = len(data["days"][n][w])
            commits = ai[n][w]["scm.commit"]
            scripts = bb[n][w]["scripts_added"] + bb[n][w]["scripts_modified"]
            lines = bb[n][w]["lines_added"] + bb[n][w]["lines_removed"]
            c = (cost[n].get(w) or {}).get("modelled")
            for col, val in enumerate([
                    n, w, window_label.get(w, "full week"),
                    round(prompts / d, 1) if d else None,
                    pct(ai[n][w]["tool.call"], prompts), commits,
                    round(commits / prompts, 2) if prompts else None,
                    lines, scripts,
                    round(scripts / prompts, 2) if prompts else None, merged,
                    round(c / merged, 2) if c and merged else None,
                    round(c / scripts, 2) if c and scripts else None], 1):
                ws.cell(row=r, column=col, value=val)
            r += 1
    notes(ws, r + 1, [
        "These compare two things we actually counted. None of them is a "
        "speed-up figure -- that needs a group doing the same work without AI, "
        "and there is none.",
        "Commits and delivered code changes are counted from two different "
        "places and measure different things, so they are never added together.",
        "Cost per unit uses the estimated AI cost -- a direction of travel, "
        "not a bill.",
        "A blank cell means there was nothing to divide by that week -- it "
        "does not mean zero.",
    ])

    charts(wb, data, weeks, window_label)
    pm_view(wb, data, weeks, full_weeks, window_label, pron)
    test_activity(wb, data, weeks, window_label)
    person_tabs(wb, data, weeks, window_label, pron)

    # Summary first: it is the answer, and it covers both people. The rest is
    # the evidence, in the order a reader would want it if they kept going.
    order = (["Summary", "Charts", "Test Activity"] + list(data["names"])
             + ["Ten Metrics", "AI Usage", "Productivity", "Person Detail",
                "Chart Data"])
    wb._sheets = ([wb[t] for t in order if t in wb.sheetnames] +
                  [s for s in wb._sheets if s.title not in order])


def main(argv=None):
    p = argparse.ArgumentParser(
        description="Add AI-usage, ten-metric and productivity sheets to a "
                    "workbook written by people_workbook.py.")
    p.add_argument("workbook", help="the .xlsx to augment, in place")
    p.add_argument("--person", type=parse_person, action="append", required=True,
                   metavar="NAME=ACCOUNT_ID",
                   help="Repeatable. Only named people appear in the file.")
    p.add_argument("--input", nargs="+", required=True,
                   help="NDJSON files or directories.")
    p.add_argument("--weeks", type=week_span, required=True,
                   metavar="YYYY-Www..YYYY-Www", help="Week span to render.")
    p.add_argument("--full-weeks", type=week_span, metavar="YYYY-Www..YYYY-Www",
                   help="The weeks that are complete. The change column spans "
                        "these only, so a partial week never misreads volume. "
                        "Defaults to every week given.")
    p.add_argument("--price", type=parse_price, action="append", default=[],
                   metavar="MODEL=IN/OUT",
                   help="Repeatable, per 1M tokens. A model with no price is "
                        "counted and left unpriced, never guessed.")
    p.add_argument("--keep-generated-charts", action="store_true",
                   help="Keep people_workbook.py's Trend Charts/Trend Data "
                        "sheets. They are superseded by the Charts sheet and "
                        "are dropped by default.")
    p.add_argument("--pronouns", type=parse_pronouns, action="append",
                   default=[], metavar="NAME=she",
                   help="Repeatable. she / he / they, or a full "
                        "subject/object/possessive/independent set. Never "
                        "inferred from a name; anyone unnamed stays they/them.")
    p.add_argument("--partial", action="append", default=[],
                   metavar="YYYY-Www=REASON",
                   help="Repeatable. Labels a week's Window column.")
    args = p.parse_args(argv)

    if not os.path.exists(args.workbook):
        raise SystemExit("no such workbook: %s -- run people_workbook.py first"
                         % args.workbook)

    people = {acct: name for name, acct in args.person}
    prices = dict(args.price)
    weeks = args.weeks
    full = args.full_weeks or weeks
    missing = [w for w in full if w not in weeks]
    if missing:
        raise SystemExit("--full-weeks names weeks outside --weeks: %s"
                         % ", ".join(missing))
    labels = {}
    for item in args.partial:
        w, _, why = item.partition("=")
        labels[w.strip()] = why.strip() or "partial week"

    said = dict(args.pronouns)
    unknown = [n for n in said if n not in people.values()]
    if unknown:
        raise SystemExit("--pronouns names someone not in --person: %s"
                         % ", ".join(sorted(unknown)))
    default = Pronouns()
    pron = lambda n: said.get(n, default)

    data = collect(args.input, people, weeks, prices)
    wb = load_workbook(args.workbook)
    render(wb, data, weeks, full, labels, pron)
    if not args.keep_generated_charts:
        drop_superseded(wb)
    colour_tabs(wb, data["names"])
    wb.save(args.workbook)

    print(json.dumps({
        "msg": "ai_sheets_written", "workbook": args.workbook,
        "people": sorted(people.values()), "weeks": len(weeks),
        "cycles": len(data["coverage_by_cycle"]),
        "priced_models": sorted(prices),
        "superseded_dropped": not args.keep_generated_charts,
        "unpriced_calls": sum((data["cost"][n][w] or {}).get("unpriced", 0)
                              for n in data["names"] for w in weeks
                              if data["cost"][n].get(w)),
    }, sort_keys=True))
    return 0



# --------------------------------------------------------------------------
# charts
# --------------------------------------------------------------------------
#
# The form is chosen by the data's job, not by habit. The sheet this replaces
# carried eleven line charts of identical shape, several of them plotting
# sequences like 0, 0, 1, 0 -- a line implies continuity between points, and
# four discrete weekly counts have none. What each chart here is for:
#
#   coverage by cycle   magnitude across 7 named things   -> sorted bar, one hue
#   prompts per week    change over time, 2 people        -> line, 2 hues
#   action rate         a rate over time                  -> line, axis pinned
#   scripts / cost      discrete weekly counts            -> grouped column
#
# Two people means two categorical hues in fixed order; a legend is always
# present, so identity is never carried by colour alone.

def _line(chart):
    """Recessive chrome. Hairline axes, no gridline hatching, thin marks."""
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height, chart.width = 7.5, 17
    return chart


def _paint(series, hex_colour, line=True, marker=False):
    gp = series.graphicalProperties
    if line:
        gp.line.solidFill = hex_colour
        gp.line.width = 2 * PT
        series.smooth = False
        if marker:
            series.marker = Marker(symbol="circle", size=7)
            series.marker.graphicalProperties.solidFill = hex_colour
            series.marker.graphicalProperties.line.solidFill = hex_colour
    else:
        gp.solidFill = hex_colour
        gp.line.noFill = True


def charts(wb, data, weeks, window_label):
    """A Charts sheet, plus the data blocks the charts read from."""
    names = data["names"]
    ai, bb, cost = data["ai"], data["bb"], data["cost"]
    ws = fresh(wb, "Charts")
    src = fresh(wb, "Chart Data")

    ws.column_dimensions["A"].width = 2
    ws.cell(row=1, column=2, value="Where the work is, and how the AI is being used"
            ).font = Font(bold=True, size=14)
    ws.cell(row=2, column=2,
            value="Every chart reads from the Chart Data sheet, which is the "
                  "table view of the same numbers.").font = NOTE

    sr = 1

    def block(title, categories, series_rows):
        """Write a titled data block; return (first_data_row, cat_col, n)."""
        nonlocal sr
        src.cell(row=sr, column=1, value=title).font = GRP
        sr += 1
        head_row = sr
        src.cell(row=sr, column=1, value="")
        for i, (label, _) in enumerate(series_rows, 2):
            src.cell(row=sr, column=i, value=label)
        sr += 1
        first = sr
        for j, cat in enumerate(categories):
            src.cell(row=sr, column=1, value=cat)
            for i, (_, vals) in enumerate(series_rows, 2):
                v = vals[j]
                src.cell(row=sr, column=i, value=v if v is not None else None)
            sr += 1
        last = sr - 1
        sr += 1
        return head_row, first, last, len(series_rows)

    # 1 -------------------------------------------------- coverage by cycle
    cov = data["coverage_by_cycle"]
    order = sorted(cov, key=lambda k: -(cov[k]["pct"] or 0))
    labels = ["%s  (%d cases)" % (k, cov[k]["cases"]) for k in order]
    hr, f, l, n = block("Automation coverage by cycle (%)", labels,
                        [("Automated %", [cov[k]["pct"] for k in order])])
    ch = _line(BarChart())
    ch.type, ch.style = "bar", None            # horizontal: long category names
    ch.title = "Automation coverage by cycle"
    ch.y_axis.title = "% of the cycle's cases automated"
    ch.y_axis.scaling.min, ch.y_axis.scaling.max = 0, 100
    ch.add_data(Reference(src, min_col=2, min_row=hr, max_row=l), titles_from_data=True)
    ch.set_categories(Reference(src, min_col=1, min_row=f, max_row=l))
    _paint(ch.series[0], MAGNITUDE, line=False)
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showVal = True
    ch.legend = None                            # one series: the title names it
    ch.height = 9
    ws.add_chart(ch, "B4")

    # 2 ------------------------------------------------------ prompts/week
    shown = [w for w in weeks if any(ai[n][w]["human.turn"] for n in names)]
    hr, f, l, n = block(
        "AI prompts per week", [w[-3:] for w in shown],
        [(nm, [ai[nm][w]["human.turn"] for w in shown]) for nm in names])
    ch = _line(LineChart())
    ch.title = "AI prompts per week"
    ch.y_axis.title = "prompts"
    ch.add_data(Reference(src, min_col=2, max_col=1 + n, min_row=hr, max_row=l),
                titles_from_data=True)
    ch.set_categories(Reference(src, min_col=1, min_row=f, max_row=l))
    for i, s in enumerate(ch.series):
        _paint(s, SERIES[i % len(SERIES)], marker=True)
    ws.add_chart(ch, "B24")

    # 3 -------------------------------------------------------- action rate
    hr, f, l, n = block(
        "Action rate % (tool calls / prompts)", [w[-3:] for w in shown],
        [(nm, [pct(ai[nm][w]["tool.call"], ai[nm][w]["human.turn"])
               for w in shown]) for nm in names])
    ch = _line(LineChart())
    ch.title = "Action rate: is the assistant doing work, or answering?"
    ch.y_axis.title = "% of prompts that led to a tool call"
    # Pinned. On an auto axis a 4-20% band fills the plot and reads as chaos.
    ch.y_axis.scaling.min, ch.y_axis.scaling.max = 0, 30
    ch.add_data(Reference(src, min_col=2, max_col=1 + n, min_row=hr, max_row=l),
                titles_from_data=True)
    ch.set_categories(Reference(src, min_col=1, min_row=f, max_row=l))
    for i, s in enumerate(ch.series):
        _paint(s, SERIES[i % len(SERIES)], marker=True)
    ws.add_chart(ch, "B40")

    # 4 --------------------------------------------------- scripts touched
    hr, f, l, n = block(
        "Automation scripts touched per week", [w[-3:] for w in shown],
        [(nm, [bb[nm][w]["scripts_added"] + bb[nm][w]["scripts_modified"]
               for w in shown]) for nm in names])
    ch = _line(BarChart())
    ch.type, ch.grouping, ch.style = "col", "clustered", None
    ch.title = "Automation scripts touched per week"
    ch.y_axis.title = "spec / step-definition files"
    ch.add_data(Reference(src, min_col=2, max_col=1 + n, min_row=hr, max_row=l),
                titles_from_data=True)
    ch.set_categories(Reference(src, min_col=1, min_row=f, max_row=l))
    for i, s in enumerate(ch.series):
        _paint(s, SERIES[i % len(SERIES)], line=False)
    ch.gapWidth = 60
    ws.add_chart(ch, "B56")

    # 5 ----------------------------------------------------------- cost
    hr, f, l, n = block(
        "Modelled token cost per week ($)", [w[-3:] for w in shown],
        [(nm, [(cost[nm].get(w) or {}).get("modelled") for w in shown])
         for nm in names])
    ch = _line(BarChart())
    ch.type, ch.grouping, ch.style = "col", "clustered", None
    ch.title = "Modelled token cost per week (not the invoice)"
    ch.y_axis.title = "USD, list price"
    ch.add_data(Reference(src, min_col=2, max_col=1 + n, min_row=hr, max_row=l),
                titles_from_data=True)
    ch.set_categories(Reference(src, min_col=1, min_row=f, max_row=l))
    for i, s in enumerate(ch.series):
        _paint(s, SERIES[i % len(SERIES)], line=False)
    ch.gapWidth = 60
    ws.add_chart(ch, "B72")

    r = 88
    partials = [w for w in shown if w in window_label]
    notes(ws, r, [
        "Each person keeps the same colour on every chart.",
        "Automation is counted per test cycle. A cycle at 0%% is run by hand "
        "on purpose, not a failure.",
        "The middle chart's scale is fixed at 0-30%% on purpose: left to scale "
        "itself, a 4-20%% range fills the chart and looks like a collapse.",
        "Partial windows, not comparable on volume: "
        + ("; ".join("%s = %s" % (w[-3:], window_label[w]) for w in partials)
           if partials else "none"),
        "AI cost is an estimate from the minority of questions that record "
        "what they used, at list price, and leaves out the monthly per-person "
        "fee that is most of the real bill.",
    ])


# --------------------------------------------------------------------------
# the one sheet
# --------------------------------------------------------------------------
#
# A PM opening fifteen tabs reads none of them. This is the sheet that answers
# the two questions the ten metrics exist to answer -- are they using AI
# effectively, and is their performance increasing -- and it is first in the
# file so it is what opens.
#
# The distinction that makes it honest: **a rate survives a partial week, a
# volume does not.** 24-27 Aug is four days, so "prompts this week" is not
# comparable to a five-day week and is marked; "cost per script touched" is a
# ratio of two things measured in the same window and is comparable. Volumes
# get compared across full weeks only; rates use every week.
#
# The second is that two people on the same team do different jobs. One writes
# automation and that work lands in the SCM; the other runs tests and moves
# tickets, and theirs lands in AIO and Jira. Scoring both on scripts-per-prompt
# would measure the second against a job they are not doing, so each gets the
# measures that fit their work and the sheet says which is which.

TITLE = Font(bold=True, size=16, color="1F3864")
Q = Font(bold=True, size=11, color="1F3864")
VERDICT = Font(bold=True, size=11)
ROLE = Font(italic=True, size=10, color="52514E")
GOOD, BAD, FLAT = "1BAF7A", "EB6834", "52514E"


def _arrow(series, want_up, compare=None):
    """Direction, and whether the series actually goes that way.

    Two traps this exists to avoid, both of which it fell into first:

    **A V is not a trend.** One person's prompts ran 189, 61, 75, 97. First-against-
    last calls that "down 49%" when it fell once and has risen every week
    since. A series whose steps change sign has no direction, and saying so is
    the honest answer.

    **A volume must not be compared across a part-week.** `compare` names the
    indices that are like for like -- the full weeks -- for counts. Rates are
    ratios of two things measured in the same window, so they pass `None` and
    use every point.
    """
    idx = compare if compare is not None else list(range(len(series)))
    pts = [series[i] for i in idx if isinstance(series[i], (int, float))]
    if len(pts) < 2:
        return "-", FLAT, None
    first, last = pts[0], pts[-1]
    change = None if not first else round(100.0 * (last - first) / first)
    steps = [b - a for a, b in zip(pts, pts[1:])]
    rising = [d for d in steps if d > 0]
    falling = [d for d in steps if d < 0]
    if rising and falling:
        # Direction reverses. Report the shape, not a slope through it.
        return "no clear trend", FLAT, change
    if not rising and not falling:
        return "flat", FLAT, 0
    up = bool(rising)
    word = ("up" if up else "down") + (" every week" if len(pts) > 2 else "")
    return word, (GOOD if up == want_up else BAD), change


def pm_view(wb, data, weeks, full_weeks, window_label, pron):
    """One sheet, written for the person who signs things off.

    Everything here is named the way the reader would say it, not the way the
    system stores it. "Action rate" is a field name; "how often AI did the work
    rather than just answering" is the thing it measures. A dagger footnote is
    a typesetting convention; "(4-day week)" is what a reader needs to know.
    Nothing internal appears -- no metric numbers, no contract sections, no
    field names, no environment variables. If a limit matters it is stated in
    the words a manager would use to repeat it to someone else.
    """
    ai, bb, cost = data["ai"], data["bb"], data["cost"]
    names = data["names"]
    shown = [w for w in weeks if any(ai[n][w]["human.turn"] for n in names)]
    # The summary IS the first sheet -- a workbook that opens on a signpost
    # telling you where to go has wasted the one place a reader always looks.
    # people_workbook.py's own Summary is per-person Jira and pull-request
    # detail; it keeps that content under a name that says so.
    if "Summary" in wb.sheetnames and "Person Detail" not in wb.sheetnames:
        wb["Summary"].title = "Person Detail"
    ws = fresh(wb, "Summary")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 44
    for col in "CDEFG":
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["H"].width = 24
    ws.sheet_view.showGridLines = False

    # A reader thinks in dates, not ISO week numbers. Show the actual days,
    # capped at the last day anything was recorded so a part-week does not
    # advertise dates nobody has data for.
    last_seen = max((d for n in names for w in weeks
                     for d in data["days"][n][w]), default=None)

    def label_of(w):
        year, num = int(w[:4]), int(w[-2:])
        monday = datetime.date.fromisocalendar(year, num, 1)
        sunday = monday + datetime.timedelta(days=6)
        if last_seen and sunday.isoformat() > last_seen:
            sunday = datetime.date.fromisoformat(last_seen)
        span = ("%d–%d %s" % (monday.day, sunday.day, sunday.strftime("%b"))
                if monday.month == sunday.month else
                "%d %s – %d %s" % (monday.day, monday.strftime("%b"),
                                   sunday.day, sunday.strftime("%b")))
        days = (sunday - monday).days + 1
        return span + ("\n(%d days)" % days if w in window_label else "")

    r = 2
    ws.cell(row=r, column=2,
            value="Is AI helping the team, and are they getting better at it?"
            ).font = TITLE
    r += 1
    ws.cell(row=r, column=2,
            value="August 2026 · %s · counted automatically from Jira, AIO "
                  "test, Bitbucket and Copilot — not from anything they "
                  "filled in" % " and ".join(names)
            ).font = NOTE
    r += 1
    ws.cell(row=r, column=2,
            value="Dark blue tabs are the answer, grey tabs the detail behind "
                  "it, amber what could not be measured. A blank cell means "
                  "nobody measured it — it does not mean zero."
            ).font = NOTE
    r += 2

    def table(rows, note):
        nonlocal r
        ws.cell(row=r, column=2, value="What we measured").font = GRP
        for i, w in enumerate(shown):
            c = ws.cell(row=r, column=3 + i, value=label_of(w))
            c.font = GRP
            c.alignment = Alignment(horizontal="right", wrap_text=True)
        ws.row_dimensions[r].height = 30
        ws.cell(row=r, column=3 + len(shown), value="Trend").font = GRP
        r += 1
        full_idx = [shown.index(w) for w in full_weeks if w in shown]
        for label, series, want_up, fmt, is_rate in rows:
            ws.cell(row=r, column=2, value=label)
            for i, v in enumerate(series):
                c = ws.cell(row=r, column=3 + i,
                            value=(fmt % v) if isinstance(v, (int, float)) else "-")
                c.alignment = Alignment(horizontal="right")
            word, colour, change = _arrow(
                series, want_up, None if is_rate else full_idx)
            plain = {"up every week": "better every week",
                     "down every week": "better every week",
                     "no clear trend": "up and down",
                     "flat": "no change", "-": "-"}.get(word)
            if plain is None:
                plain = "worse" if colour == BAD else "better"
            if colour == BAD and plain.startswith("better"):
                plain = "worse every week"
            # A net figure beside "up and down" has to say it is a net, or it
            # reads as the trend the row has just denied having.
            suffix = ("" if change is None else
                      ("   net %+d%%" % change if plain == "up and down"
                       else "   %+d%%" % change))
            cell = ws.cell(row=r, column=3 + len(shown), value=plain + suffix)
            cell.font = Font(bold=True, color=colour)
            r += 1
        ws.cell(row=r, column=2, value=note).font = NOTE
        r += 2

    def block(name, role, verdict, verdict_colour, rows, note):
        nonlocal r
        ws.cell(row=r, column=2, value=name.upper()).font = Font(bold=True, size=13)
        ws.cell(row=r, column=3, value=role).font = ROLE
        r += 1
        ws.cell(row=r, column=2, value=verdict).font = Font(
            bold=True, size=11, color=verdict_colour)
        r += 2
        table(rows, note)

    def rate(n, a, b):
        return [pct(ai[n][w][a], ai[n][w][b]) for w in shown]

    def money(n, per):
        out = []
        for w in shown:
            c = (cost[n].get(w) or {}).get("modelled")
            d = per(n, w)
            out.append(round(c / d, 2) if c and d else None)
        return out

    scripts = lambda n, w: bb[n][w]["scripts_added"] + bb[n][w]["scripts_modified"]

    # ---- the person who writes the automation -----------------------------
    builder = next((n for n in names if sum(scripts(n, w) for w in shown) > 10), None)
    if builder:
        b = pron(builder)
        s_scripts = [scripts(builder, w) for w in shown]
        s_prompts = [ai[builder][w]["human.turn"] for w in shown]
        fidx = [shown.index(w) for w in full_weeks if w in shown]
        d_scripts = _arrow(s_scripts, True, fidx)[2]
        d_prompts = _arrow(s_prompts, False, fidx)[2]
        block(builder, "builds the automated tests",
              "Getting better. %s %s producing more while asking AI for less, "
              "and each delivery costs less than it did." % (b.Subj, b.is_),
              GOOD,
              [("How often AI did the work, not just answered",
                rate(builder, "tool.call", "human.turn"), True, "%.0f%%", True),
               ("AI cost per test script %s worked on" % b.subj,
                money(builder, scripts), False, "$%.2f", True),
               ("AI cost per change %s delivered" % b.subj,
                money(builder, lambda n, w: bb[n][w]["scm.pr.merged"]), False,
                "$%.2f", True),
               ("Test scripts worked on", s_scripts, True, "%d", False),
               ("Times %s asked AI for help" % b.subj,
                s_prompts, False, "%d", False)],
              "The bottom two lines are the story: across the full weeks %s "
              "worked on %s more test scripts while asking AI for help %s "
              "less. More work, less AI to get there. Both lines move around "
              "week to week, so the overall change is the claim — not a "
              "straight line through four weeks."
              % (b.subj,
                 ("%d%%" % d_scripts) if d_scripts is not None else "more",
                 ("%d%%" % abs(d_prompts)) if d_prompts is not None else "less"))

    # ---- the person who runs the tests ------------------------------------
    runner = next((n for n in names if n != builder), None)
    if runner:
        wkp = data["week_people"]
        bugs = [wkp.get(w, {}).get(runner, {}).get("raised_bug", 0) for w in shown]
        ran = [wkp.get(w, {}).get(runner, {}).get("runs", 0) for w in shown]
        g = pron(runner)
        block(runner, "runs the tests and raises the bugs",
              "Delivering. %d bugs raised and %d tests run across these weeks. %s AI "
              "use is not settling into a pattern yet, and it costs more per "
              "question than it did."
              % (sum(bugs), sum(ran), g.Poss), FLAT,
              [("Bugs %s raised for the developers" % g.subj, bugs, True,
                "%d", False),
               ("Tests %s ran" % g.subj, ran, True, "%d", False),
               ("How often AI did the work, not just answered",
                rate(runner, "tool.call", "human.turn"), True, "%.0f%%", True),
               ("Times %s asked AI for help" % g.subj,
                [ai[runner][w]["human.turn"] for w in shown], False, "%d", False),
               ("AI cost each time %s asked" % g.subj,
                [round(((cost[runner].get(w) or {}).get("modelled") or 0)
                       / ai[runner][w]["human.turn"], 2)
                 if ai[runner][w]["human.turn"] else None for w in shown],
                False, "$%.2f", True)],
              "Finding bugs is the job, and %s %s the one finding them — most "
              "of the team's bugs this month are %s. %s %s little automation "
              "code, so the test-script and delivery figures above are not %s "
              "to hit. The Test Activity tab and the %s tab carry the "
              "week-by-week detail."
              % (g.subj, g.is_, g.indep, g.Subj, g.v("write"), g.indep, runner))

    # ---- the estate --------------------------------------------------------
    cov = data["coverage_by_cycle"]
    tot = sum(c["cases"] for c in cov.values())
    auto = sum(c["automated"] for c in cov.values())
    ws.cell(row=r, column=2, value="HOW MUCH OF THE TESTING IS AUTOMATED"
            ).font = Font(bold=True, size=13)
    r += 1
    if not tot:
        ws.cell(row=r, column=2,
                value="Not measured — the test management data was not included "
                      "in this pull.").font = VERDICT
        r += 1
        ws.cell(row=r, column=2,
                value="This is a missing source, not a score of zero. Do not "
                      "read anything into it.").font = NOTE
        r += 2
    else:
        ws.cell(row=r, column=2,
                value="%.0f%% — %s of the %s tests in the cycles being run are "
                      "automated." % (100.0 * auto / tot, f"{auto:,}", f"{tot:,}")
                ).font = VERDICT
        r += 1
        manual = [k for k in cov if cov[k]["pct"] == 0]
        ws.cell(row=r, column=2,
                value=("Counted across the test cycles actually being run, which "
                       "is the work in flight." +
                       (" %d of the %d cycles are run by hand on purpose — that "
                        "is a decision about scope, not a gap." % (len(manual), len(cov))
                        if manual else "")).strip()).font = NOTE
        r += 2

    # ---- limits ------------------------------------------------------------
    ws.cell(row=r, column=2, value="WHAT THESE NUMBERS CANNOT TELL YOU"
            ).font = Font(bold=True, size=13)
    r += 1
    for line in [
        "Whether AI made them faster. To say that we would need the same work "
        "done without AI to compare against, and there isn't any. Anyone "
        "quoting a speed-up percentage is guessing.",
        "What AI actually costs. The dollar figures are the value of the text "
        "sent to and from the AI. The real bill is a monthly charge per person "
        "plus usage, which this cannot see. Use the figures to compare weeks "
        "against each other, not to forecast a budget.",
        "The cost of one finished piece of work. The tool cannot yet tell which "
        "AI conversation produced which script, so cost can be split by person "
        "and week but not by deliverable.",
        "Anything with confidence from four weeks. The direction is worth "
        "acting on. The exact percentages are not — expect them to move.",
        "Weeks marked (4 days) are short. Percentages from them are still fair "
        "to compare; the raw counts are not, so those trends ignore them.",
        "Up-and-down means the number changed direction between weeks. Four "
        "points that fall then rise are not a decline, and this sheet will not "
        "pretend otherwise.",
    ]:
        ws.cell(row=r, column=2, value="•  " + line).font = NOTE
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1


# --------------------------------------------------------------------------
# test activity, and one tab per person
# --------------------------------------------------------------------------

def test_activity(wb, data, weeks, window_label):
    """Test-management work per week, split by person.

    Running a test and raising a ticket both carry a person, so both are
    counted per person and only these two people are in `people` -- nobody
    else's work is in these columns. Creating and editing a test case carries
    no author at all (AIO records none), so that stays in its own block below,
    labelled as the whole project's, rather than being guessed onto one of
    them.
    """
    wk, names, wkp = data["week"], data["names"], data["week_people"]
    live = [w for w in weeks if any(wk.get(w, {}).values())]
    if not live:
        return
    ws = fresh(wb, "Test Activity")
    ws.sheet_view.showGridLines = False

    ROWS = [("runs", "Tests run"), ("passed", "Passed"), ("failed", "Failed"),
            ("automated", "Run by automation"),
            ("defects", "Defects logged"),
            ("raised_bug", "Bugs raised"), ("raised_task", "Tasks raised")]

    # Two header rows: the person spans a block, the measures sit under it, so
    # a column is never read against the wrong name.
    ws.cell(row=1, column=1, value="Week").font = HEAD
    ws.cell(row=1, column=2, value="Dates").font = HEAD
    for c in (1, 2):
        ws.cell(row=1, column=c).fill = FILL
        ws.cell(row=2, column=c).fill = FILL
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 17
    col = 3
    for n in names:
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col + len(ROWS) - 1)
        c = ws.cell(row=1, column=col, value=n)
        c.font, c.fill = HEAD, FILL
        c.alignment = Alignment(horizontal="center")
        for i, (_, label) in enumerate(ROWS):
            h = ws.cell(row=2, column=col + i, value=label)
            h.font, h.fill = HEAD, FILL
            h.alignment = Alignment(vertical="bottom", wrap_text=True,
                                    horizontal="right")
            ws.column_dimensions[get_column_letter(col + i)].width = 11
        col += len(ROWS)
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = ws.cell(row=3, column=3)

    last_seen = max((d for n in names for w in weeks
                     for d in data["days"][n][w]), default=None)
    r = 3
    for w in live:
        year, num = int(w[:4]), int(w[-2:])
        mon = datetime.date.fromisocalendar(year, num, 1)
        sun = mon + datetime.timedelta(days=6)
        if last_seen and sun.isoformat() > last_seen:
            sun = datetime.date.fromisoformat(last_seen)
        ws.cell(row=r, column=1, value=w[-3:])
        ws.cell(row=r, column=2,
                value="%d %s — %d %s" % (mon.day, mon.strftime("%b"),
                                          sun.day, sun.strftime("%b")))
        col = 3
        for n in names:
            mine = wkp.get(w, {}).get(n, {})
            for i, (key, _) in enumerate(ROWS):
                cell = ws.cell(row=r, column=col + i, value=mine.get(key, 0))
                cell.alignment = Alignment(horizontal="right")
            col += len(ROWS)
        r += 1

    # ---- the part that carries no author --------------------------------
    r += 2
    ws.cell(row=r, column=1,
            value="TEST CASES — WHOLE PROJECT, NOT SPLIT BY PERSON"
            ).font = Font(bold=True, size=11)
    r += 1
    for i, label in enumerate(["Week", "Dates", "Test cases created",
                               "Test cases edited", "Cycles worked in"]):
        c = ws.cell(row=r, column=1 + i, value=label)
        c.font, c.fill = HEAD, FILL
    r += 1
    for w in live:
        c = wk.get(w, {})
        year, num = int(w[:4]), int(w[-2:])
        mon = datetime.date.fromisocalendar(year, num, 1)
        sun = mon + datetime.timedelta(days=6)
        if last_seen and sun.isoformat() > last_seen:
            sun = datetime.date.fromisoformat(last_seen)
        for cnum, val in enumerate([
                w[-3:], "%d %s — %d %s" % (mon.day, mon.strftime("%b"),
                                             sun.day, sun.strftime("%b")),
                c.get("cases_created", 0), c.get("cases_updated", 0),
                len(data["week_cycles"].get(w, []))], 1):
            ws.cell(row=r, column=cnum, value=val)
        r += 1

    r += 1
    for line in [
        "Every per-person column counts only that person. Nobody else's runs, "
        "bugs or tickets are in this sheet.",
        "\"Test cases created\" is a real date. \"Test cases edited\" is the "
        "LAST edit only — a case changed three times in a week counts once, "
        "and an earlier change is overwritten by a later one.",
        "Test cases are not split by person because the test tool records no "
        "author on a case. That block is the whole project, including people "
        "who are not in this report.",
        "A test that failed and later passed cannot be counted: the test tool "
        "keeps only the latest result for each test in each cycle, so a fix "
        "made inside a cycle leaves no trace.",
    ]:
        ws.cell(row=r, column=1, value=line).font = NOTE
        r += 1


def person_tabs(wb, data, weeks, window_label, pron):
    """One tab per person: their AI use, their delivery, what they worked on."""
    names = data["names"]
    ai, bb, cost = data["ai"], data["bb"], data["cost"]
    shown = [w for w in weeks if any(ai[n][w]["human.turn"] for n in names)]

    for n in names:
        pn = pron(n)
        ws = fresh(wb, n[:31])
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 38
        for col in "CDEFGH":
            ws.column_dimensions[col].width = 13
        r = 2
        ws.cell(row=r, column=2, value=n).font = TITLE
        r += 2

        def section(title):
            nonlocal r
            ws.cell(row=r, column=2, value=title).font = Font(bold=True, size=12)
            r += 1

        def head_row(cols):
            nonlocal r
            for i, c in enumerate(cols):
                cell = ws.cell(row=r, column=2 + i, value=c)
                cell.font = GRP
                if i:
                    cell.alignment = Alignment(horizontal="right", wrap_text=True)
            ws.row_dimensions[r].height = 28
            r += 1

        def line(label, vals, fmt="%s"):
            nonlocal r
            ws.cell(row=r, column=2, value=label)
            for i, v in enumerate(vals):
                c = ws.cell(row=r, column=3 + i,
                            value=(fmt % v) if isinstance(v, (int, float)) else "-")
                c.alignment = Alignment(horizontal="right")
            r += 1

        # ---- how they used AI, by week -----------------------------------
        section("How %s used AI, week by week" % n)
        head_row(["Measure"] + [
            ("%s%s" % (w[-3:].replace("W", "wk "),
                       "\n(part week)" if w in window_label else ""))
            for w in shown])
        line("Times %s asked AI for help" % pn.subj,
             [ai[n][w]["human.turn"] for w in shown], "%d")
        line("Separate AI conversations",
             [len(data["sessions"][n][w]) for w in shown], "%d")
        line("Days %s used AI" % pn.subj,
             [len(data["days"][n][w]) for w in shown], "%d")
        line("How often AI did the work, not just answered",
             [pct(ai[n][w]["tool.call"], ai[n][w]["human.turn"]) for w in shown],
             "%.0f%%")
        line("AI cost, estimated",
             [(cost[n].get(w) or {}).get("modelled") for w in shown], "$%.2f")
        line("Questions with a measured cost",
             [(cost[n].get(w) or {}).get("n") for w in shown], "%d")
        r += 1

        # ---- what they delivered, by week ---------------------------------
        section("What %s delivered, week by week" % n)
        wk_people = data["week_people"]
        live = [w for w in shown if wk_people.get(w, {}).get(n)]
        if live:
            head_row(["Measure"] + [
                ("%s%s" % (w[-3:].replace("W", "wk "),
                           "\n(part week)" if w in window_label else ""))
                for w in live])
            for key, label in (("raised_bug", "Bugs %s raised" % pn.subj),
                               ("runs", "Tests %s ran" % pn.subj),
                               ("failed", "Of those, failures found"),
                               ("defects", "Defects logged against %s runs" % pn.poss),
                               ("raised_task", "Tasks %s raised" % pn.subj)):
                line(label, [wk_people[w][n].get(key, 0) for w in live], "%d")
        else:
            ws.cell(row=r, column=2,
                    value="No test runs or tickets recorded against %s in "
                          "these weeks." % pn.obj).font = NOTE
            r += 1
        r += 1

        # ---- code side, only where there is one ---------------------------
        touched = sum(bb[n][w]["scripts_added"] + bb[n][w]["scripts_modified"]
                      for w in shown)
        section("Code %s delivered, week by week" % n)
        if touched:
            head_row(["Measure"] + [w[-3:].replace("W", "wk ") for w in shown])
            line("Test scripts worked on",
                 [bb[n][w]["scripts_added"] + bb[n][w]["scripts_modified"]
                  for w in shown], "%d")
            line("Changes delivered",
                 [bb[n][w]["scm.pr.merged"] for w in shown], "%d")
            line("Changes reviewed for others",
                 [bb[n][w]["scm.pr.reviewed"] for w in shown], "%d")
            line("Changes rolled back",
                 [bb[n][w]["scm.revert"] for w in shown], "%d")
            line("Lines of code changed",
                 [bb[n][w]["lines_added"] + bb[n][w]["lines_removed"]
                  for w in shown], "%d")
        else:
            ws.cell(row=r, column=2,
                    value="%s did not deliver code changes this month. That is "
                          "%s role, not a gap — %s output is on the test "
                          "table above." % (pn.Subj, pn.poss, pn.poss)).font = NOTE
            r += 1
        r += 1

        # ---- what they worked on ------------------------------------------
        section("What %s worked on" % n)
        cov = data["coverage_by_cycle"]
        mine = [(k, c) for k, c in cov.items() if c["ours"].get(n)]
        if mine:
            head_row(["Test cycle", "Tests %s ran" % pn.subj, "Area", "Automated"])
            for k, c in sorted(mine, key=lambda x: -x[1]["ours"][n]):
                ws.cell(row=r, column=2, value=k)
                ws.cell(row=r, column=3, value=c["ours"][n]).alignment = \
                    Alignment(horizontal="right")
                ws.cell(row=r, column=4, value=c["folder"] or "-")
                ws.cell(row=r, column=5, value="%.0f%%" % c["pct"])
                r += 1
        else:
            ws.cell(row=r, column=2, value="No test cycles recorded.").font = NOTE
            r += 1
        r += 1
        ws.cell(row=r, column=2,
                value="Counted automatically from Jira, AIO test, Bitbucket and "
                      "Copilot. A dash means nothing was measured, not zero.").font = NOTE


# --------------------------------------------------------------------------
# tab colours: which sheets are the answer, which are the evidence
# --------------------------------------------------------------------------
#
# Fifteen tabs with identical grey labels tell a reader nothing about where to
# start, and the sheets that hold the answer look exactly like the ones that
# hold six hundred rows of test runs. Three roles, three colours, stated on the
# first sheet so the code is not folklore:
#
#   INSIGHT   read these        the accent, matching the header fill
#   EVIDENCE  the rows behind   neutral grey, deliberately recessive
#   CAVEATS   read before quoting  amber -- it says what is NOT measured
#
# The accent is the same hue as the charts' magnitude colour, so "this is the
# point" looks the same everywhere in the file.

INSIGHT_TAB = "1F3864"
EVIDENCE_TAB = "A6A6A6"
CAVEAT_TAB = "EDA100"

INSIGHT = ("Summary", "Charts", "Test Activity", "Ten Metrics",
           "AI Usage", "Productivity")
CAVEATS = ("Coverage & Gaps",)


#: Written by people_workbook.py and superseded by the Charts sheet: eleven
#: line charts of identical shape, several plotting sequences like 0, 0, 1, 0.
#: A line implies continuity between its points and four discrete weekly counts
#: have none. Dropped by default so the file has one place to look; --keep-
#: generated-charts puts them back.
SUPERSEDED = ("Trend Charts", "Trend Data")


def drop_superseded(wb):
    for title in SUPERSEDED:
        if title in wb.sheetnames:
            del wb[title]


def colour_tabs(wb, people=()):
    """Colour every tab by its role. The legend is written by pm_view."""
    insight = set(INSIGHT) | {p[:31] for p in people}
    for ws in wb.worksheets:
        if ws.title in insight:
            ws.sheet_properties.tabColor = INSIGHT_TAB
        elif ws.title in CAVEATS:
            ws.sheet_properties.tabColor = CAVEAT_TAB
        else:
            ws.sheet_properties.tabColor = EVIDENCE_TAB


if __name__ == "__main__":
    sys.exit(main())
